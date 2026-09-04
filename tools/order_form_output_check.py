import streamlit as st

# Build marker used in Auto Detect cache keys so code updates cannot reuse
# stale detected-field selections from an older engine version.
AUTO_DETECT_ENGINE_VERSION = "2026-09-05-VISUAL-BLOCK-FAIL-ANNOTATION-07"
import pandas as pd
import fitz
import re
import unicodedata
import io
from pathlib import Path
from io import BytesIO

from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

def _apply_tool_css():
    # =========================================================
    # DARK UI
    # =========================================================

    st.markdown(
        """
        <style>

        html,
        body,
        [data-testid="stAppViewContainer"],
        [data-testid="stApp"],
        .stApp,
        .main,
        [data-testid="stMain"] {
            background-color: #0e1117 !important;
            color: #ffffff !important;
        }

        [data-testid="stHeader"] {
            background-color: #0e1117 !important;
        }

        .stApp,
        .stApp p,
        .stApp label,
        .stApp span,
        .stApp div {
            color: #ffffff;
        }

        .main-title {
            color: #ffffff !important;
            font-size: 34px;
            font-weight: 700;
            margin-top: 5px;
            margin-bottom: 4px;
        }

        .sub-title {
            color: #b8c0cc !important;
            font-size: 15px;
            margin-bottom: 30px;
        }

        .section-title {
            color: #ffffff !important;
            font-size: 20px;
            font-weight: 700;
            margin-top: 12px;
            margin-bottom: 10px;
        }

        [data-testid="stFileUploader"] {
            background-color: #161b22 !important;
            border: 1px solid #4b5563 !important;
            border-radius: 12px !important;
            padding: 8px !important;
        }

        [data-testid="stFileUploaderDropzone"] {
            background-color: #161b22 !important;
            border: 1px solid #4b5563 !important;
            border-radius: 10px !important;
        }

        [data-testid="stFileUploaderDropzoneInstructions"] {
            color: #ffffff !important;
        }

        [data-testid="stFileUploaderDropzoneInstructions"] span {
            color: #ffffff !important;
        }

        [data-testid="stFileUploader"] button {
            background-color: #111827 !important;
            color: #ffffff !important;
            border: 1px solid #6b7280 !important;
            border-radius: 8px !important;
        }

        [data-testid="stFileUploader"] button:hover {
            background-color: #1f2937 !important;
            color: #ffffff !important;
        }

        [data-baseweb="select"] > div {
            background-color: #161b22 !important;
            color: #ffffff !important;
            border: 1px solid #4b5563 !important;
            border-radius: 10px !important;
        }

        [data-baseweb="select"] input {
            color: #ffffff !important;
        }

        [data-baseweb="select"] span {
            color: #ffffff !important;
        }

        [data-baseweb="popover"] {
            background-color: #161b22 !important;
        }

        [role="option"] {
            background-color: #161b22 !important;
            color: #ffffff !important;
        }

        [role="option"]:hover {
            background-color: #263241 !important;
        }

        [data-baseweb="tag"] {
            background-color: #2563eb !important;
            color: #ffffff !important;
        }

        [data-baseweb="tag"] span {
            color: #ffffff !important;
        }

        div.stButton > button {
            background-color: #2196F3 !important;
            color: #ffffff !important;
            border: 2px solid #000000 !important;
            border-radius: 12px !important;
            font-size: 18px !important;
            font-weight: 700 !important;
            height: 54px !important;
            width: 100% !important;
            box-shadow: none !important;
        }

        div.stButton > button:hover {
            background-color: #1976D2 !important;
            color: #ffffff !important;
            border: 2px solid #000000 !important;
        }

        div.stDownloadButton > button {
            background-color: #1f2937 !important;
            color: #ffffff !important;
            border: 1px solid #6b7280 !important;
            border-radius: 10px !important;
            font-weight: 600 !important;
        }

        div.stDownloadButton > button:hover {
            background-color: #374151 !important;
            color: #ffffff !important;
        }

        [data-testid="stDataFrame"] {
            border: 1px solid #374151 !important;
            border-radius: 10px !important;
        }

        [data-testid="stMetric"] {
            background-color: #161b22 !important;
            border: 1px solid #374151 !important;
            border-radius: 10px !important;
            padding: 12px !important;
        }

        [data-testid="stMetricLabel"] {
            color: #b8c0cc !important;
        }

        [data-testid="stMetricValue"] {
            color: #ffffff !important;
        }

        hr {
            border-color: #30363d !important;
        }

        .stCaption {
            color: #9ca3af !important;
        }

        [data-testid="stAlert"] {
            border-radius: 10px !important;
        }

        [data-testid="stSpinner"] {
            color: #ffffff !important;
        }

        </style>
        """,
        unsafe_allow_html=True
    )


# =========================================================
# OPTIMIZED COMPARISON ENGINE
# =========================================================

import pandas as pd
import re
import unicodedata
import io
from PIL import Image


# =========================================================
# BASIC DATA HELPERS
# =========================================================

def is_blank_value(value):
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip() == ""


def normalize_text(text):
    if is_blank_value(text):
        return ""

    value = unicodedata.normalize("NFKC", str(text))
    value = value.casefold()

    value = value.replace("\u200b", "").replace("\ufeff", "")
    value = value.replace("’", "'").replace("`", "'")
    value = value.replace("–", "-").replace("—", "-")
    value = value.replace("\r", " ").replace("\n", " ")

    # Common PDF bullet extraction artefact.
    value = re.sub(r"(^|\s)n(?=\s)", " ", value)

    # Treat punctuation/separators as spacing differences.
    value = re.sub(r"[,.;:|/\\]+", " ", value)
    value = re.sub(r"-+", " ", value)
    value = re.sub(r"[^\w%#'\s]", " ", value, flags=re.UNICODE)

    # Apostrophe differences should not create a mismatch.
    value = value.replace("'", "")
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalize_symbol_text(text):
    if is_blank_value(text):
        return ""
    value = unicodedata.normalize("NFKC", str(text))
    value = value.replace("\u200b", "").replace("\ufeff", "")
    value = re.sub(r"\s+", "", value)
    return value.casefold()


def compact_text(text):
    return normalize_text(text).replace(" ", "")


def tokenize(text):
    value = normalize_text(text)
    return value.split() if value else []


def normalize_numeric(value):
    if is_blank_value(value):
        return None
    text = normalize_text(value)
    # Keep only a clean integer/decimal when the value is numeric.
    match = re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        number = float(match.group(0))
        if number.is_integer():
            return str(int(number))
        return str(number).rstrip("0").rstrip(".")
    except Exception:
        return None


def get_available_fields(df):
    """Only columns containing at least one populated cell are selectable."""
    result = []
    for column in df.columns:
        if df[column].map(lambda value: not is_blank_value(value)).any():
            result.append(str(column))
    return result


def load_excel(file):
    file.seek(0)
    df = pd.read_excel(file, header=0)
    df.columns = [str(column).strip() for column in df.columns]
    return df


# =========================================================
# FIELD CLASSIFICATION
# =========================================================

ADMIN_FIELD_PATTERNS = (
    r"^sr$",
    r"^sr\.?\s*no\.?$",
    r"^serial",
    r"^job\s*(?:no|number)?$",
    r"^order\s*(?:no|number|date)?$",
    r"^ticket\s*(?:no|number)?$",
    r"^created",
    r"^modified",
    r"^timestamp",
)


def get_field_region(field_name):
    original = str(field_name).casefold()
    normalized = normalize_text(field_name).replace(" ", "")

    if (
        "_en" in original
        or normalized.endswith("en")
        or "english" in normalized
    ):
        return "EN"

    if (
        "_fr" in original
        or normalized.endswith("fr")
        or "french" in normalized
        or "canada" in normalized
    ):
        return "FR"

    if (
        "_sp" in original
        or normalized.endswith("sp")
        or "spanish" in normalized
        or "espanol" in normalized
    ):
        return "SP"

    return ""


def get_field_type(field_name):
    field = normalize_text(field_name)
    compact = field.replace(" ", "").replace("_", "").replace("-", "")

    # Explicit sequence fields such as OSZ1, OSZ2 ...
    if re.fullmatch(r"osz\d+", compact):
        return "OSZ"

    if "symbol" in compact or compact in {"caremark", "caresymbol", "washsymbol"}:
        return "SYMBOL"

    if (
        compact == "rn"
        or "rnno" in compact
        or "rnnumber" in compact
        or "registrationnumber" in compact
        or "companyrn" in compact
        or compact.startswith("rn")
    ):
        return "RN"

    if (
        "sku" in compact
        or "itemcode" in compact
        or "itemnumber" in compact
        or "itemno" in compact
        or "stylecode" in compact
        or compact == "style"
        or "productcode" in compact
        or "supwsp" in compact
        or "supplier" in compact
        or "vendorid" in compact
        or "vendorcode" in compact
    ):
        return "IDENTIFIER"

    if "batch" in compact or "lotnumber" in compact or "lotno" in compact or compact == "lot":
        return "BATCH"

    if (
        "quantity" in compact
        or compact == "qty"
        or "units" in compact
        or "pieces" in compact
        or compact == "pcs"
    ):
        return "QUANTITY"

    if (
        "coo" in compact
        or "countryoforigin" in compact
        or "countryorigin" in compact
        or "madein" in compact
        or compact == "origin"
    ):
        return "COO"

    if (
        "fiber" in compact
        or "fibre" in compact
        or "fabric" in compact
        or "content" in compact
        or "composition" in compact
        or "compodsc" in compact
        or "lhcompodsc" in compact
        or "fabrication" in compact
        or "material" in compact
    ):
        return "CONTENT"

    if (
        "care" in compact
        or "wash" in compact
        or "washing" in compact
        or "laundry" in compact
        or "instruction" in compact
    ):
        return "CARE"

    if (
        "size" in compact
        or "sizeline" in compact
        or "alpha" in compact
        or "waist" in compact
        or "inseam" in compact
        or compact == "fit"
        or re.fullmatch(r"s\d+", compact)
    ):
        return "SIZE"

    if "brand" in compact:
        return "BRAND"

    if "color" in compact or "colour" in compact:
        return "COLOR"

    if "gender" in compact:
        return "GENDER"

    if "attribute" in compact or "technology" in compact or "feature" in compact:
        return "ATTRIBUTE"

    return "GENERAL"


def is_admin_field(field_name):
    normalized = normalize_text(field_name)
    return any(re.match(pattern, normalized) for pattern in ADMIN_FIELD_PATTERNS)


# =========================================================
# PDF/OCR EXTRACTION
# =========================================================

def _usable_text(text):
    if not text or not str(text).strip():
        return False
    alnum = re.sub(r"[^\w%#]", "", str(text), flags=re.UNICODE)
    return len(alnum) >= 6


def _text_quality_score(text):
    """Score extracted artwork text so OCR can be preferred over weak PDF text layers."""
    text = str(text or "")
    if not text.strip():
        return 0

    alnum = len(re.findall(r"[A-Za-z0-9]", text))
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    useful_lines = sum(1 for line in lines if len(re.sub(r"[^A-Za-z0-9%#]", "", line)) >= 2)
    numeric_runs = len(re.findall(r"(?<![A-Za-z0-9])\d+(?![A-Za-z0-9])", text))

    return alnum + useful_lines * 8 + numeric_runs * 3


def _ocr_image_with_data(image):
    """
    Multi-pass OCR for small artwork text.

    Returns:
        primary_text, primary_word_boxes, language, supplemental_text

    primary_text is reconstructed from the strongest OCR pass using physical
    word coordinates so the visual reading order is preserved. supplemental_text
    contains additional unique lines from weaker passes and is used only as
    secondary evidence for Auto Detect.
    """

    try:
        import pytesseract
        from pytesseract import Output
    except ImportError as exc:
        raise RuntimeError(
            "OCR support is not installed. Add pytesseract to requirements.txt."
        ) from exc

    try:
        from PIL import ImageOps, ImageEnhance, ImageFilter
    except Exception:
        ImageOps = ImageEnhance = ImageFilter = None

    work_image = image.convert("RGB")
    if work_image.width < 1600:
        scale = 1600 / max(1, work_image.width)
        work_image = work_image.resize(
            (int(work_image.width * scale), int(work_image.height * scale)),
            Image.Resampling.LANCZOS
        )

    variants = [("color", work_image)]
    if ImageOps is not None:
        gray = ImageOps.grayscale(work_image)
        gray = ImageOps.autocontrast(gray)
        if ImageEnhance is not None:
            gray = ImageEnhance.Contrast(gray).enhance(1.25)
        if ImageFilter is not None:
            gray = gray.filter(ImageFilter.SHARPEN)
        variants.append(("gray", gray))

    # Keep the OCR workload controlled: 3 primary passes.
    requested_passes = [
        ("color", 11, "eng"),
        ("gray", 11, "eng"),
        ("color", 6, "eng"),
    ]

    errors = []
    results = []

    for variant_name, psm, lang in requested_passes:
        variant_image = dict(variants).get(variant_name, work_image)
        try:
            data = pytesseract.image_to_data(
                variant_image,
                lang=lang,
                output_type=Output.DICT,
                config=f"--psm {psm}"
            )

            words = []
            grouped_text = {}
            conf_values = []

            for i, raw_text in enumerate(data.get("text", [])):
                word = str(raw_text or "").strip()
                if not word:
                    continue
                try:
                    conf = float(
                        data.get("conf", ["-1"] * len(data.get("text", [])))[i]
                    )
                except Exception:
                    conf = -1.0

                item = {
                    "text": word,
                    "left": int(data.get("left", [0])[i]),
                    "top": int(data.get("top", [0])[i]),
                    "width": int(data.get("width", [0])[i]),
                    "height": int(data.get("height", [0])[i]),
                    "conf": conf,
                    "block_num": int(data.get("block_num", [0])[i]),
                    "par_num": int(data.get("par_num", [0])[i]),
                    "line_num": int(data.get("line_num", [0])[i]),
                }
                words.append(item)
                if conf >= 0:
                    conf_values.append(conf)
                line_key = (
                    item["block_num"],
                    item["par_num"],
                    item["line_num"]
                )
                grouped_text.setdefault(line_key, []).append(item)

            text_lines = []
            for _key, line_words in sorted(grouped_text.items(), key=lambda pair: pair[0]):
                line_words.sort(key=lambda item: (item.get("top", 0), item.get("left", 0)))
                text_lines.append(" ".join(item["text"] for item in line_words))

            text = "\n".join(text_lines)
            if not _usable_text(text):
                continue

            avg_conf = sum(conf_values) / len(conf_values) if conf_values else 0

            # Prefer a clean reading order over a noisy OCR pass that happens
            # to contain more total characters. Artwork often contains logos,
            # barcode fragments and isolated symbols that inflate raw length.
            non_empty_lines = [line.strip() for line in text.splitlines() if line.strip()]
            junk_lines = sum(
                1
                for line in non_empty_lines
                if len(re.sub(r"[^A-Za-z0-9%#]", "", line)) <= 1
            )
            quality = (
                _text_quality_score(text)
                + (avg_conf * 0.20)
                - (junk_lines * 18)
            )

            results.append({
                "text": text,
                "words": words,
                "lang": lang,
                "quality": quality,
            })
        except Exception as exc:
            errors.append(f"{variant_name}/psm{psm}/{lang}: {exc}")

    if not results:
        if errors:
            raise RuntimeError(
                "OCR could not run. Tesseract may be missing. Details: "
                + " | ".join(errors[:4])
            )
        return "", [], "", ""

    best = max(results, key=lambda item: item["quality"])

    # Rebuild the strongest OCR pass from physical coordinates. Tesseract's
    # block/line numbering can occasionally reverse adjacent words on artwork;
    # geometry is a safer source of visual reading order.
    ordered_words = [
        word for word in best["words"]
        if str(word.get("text", "")).strip()
    ]

    primary_lines = []
    if ordered_words:
        heights = [max(1, int(word.get("height", 1))) for word in ordered_words]
        median_height = float(sorted(heights)[len(heights) // 2]) if heights else 20.0
        line_tolerance = max(10.0, median_height * 0.65)

        line_groups = []
        for word in sorted(
            ordered_words,
            key=lambda item: (
                float(item.get("top", 0)) + float(item.get("height", 0)) / 2.0,
                float(item.get("left", 0)),
            )
        ):
            center_y = (
                float(word.get("top", 0))
                + float(word.get("height", 0)) / 2.0
            )

            best_group = None
            best_distance = None
            for group in line_groups:
                distance = abs(center_y - group["center_y"])
                if distance <= line_tolerance and (
                    best_distance is None or distance < best_distance
                ):
                    best_group = group
                    best_distance = distance

            if best_group is None:
                line_groups.append({
                    "center_y": center_y,
                    "words": [word],
                })
            else:
                best_group["words"].append(word)
                best_group["center_y"] = sum(
                    float(w.get("top", 0)) + float(w.get("height", 0)) / 2.0
                    for w in best_group["words"]
                ) / len(best_group["words"])

        line_groups.sort(key=lambda group: group["center_y"])

        for group in line_groups:
            group["words"].sort(key=lambda item: float(item.get("left", 0)))
            line = " ".join(
                str(word.get("text", "")).strip()
                for word in group["words"]
                if str(word.get("text", "")).strip()
            )
            line = re.sub(r"\s+", " ", line).strip()
            if line:
                primary_lines.append(line)

    if not primary_lines:
        primary_lines = [
            re.sub(r"\s+", " ", line).strip()
            for line in best["text"].splitlines()
            if line.strip()
        ]

    seen = {normalize_text(line) for line in primary_lines if normalize_text(line)}

    # We do not append weaker OCR lines to the primary comparison text because
    # that can introduce incorrect duplicate/alternate readings. Instead,
    # return them separately for Auto Detect's secondary evidence.
    supplemental_lines = []
    for result in sorted(results, key=lambda item: item["quality"], reverse=True):
        if result is best:
            continue
        for line in result["text"].splitlines():
            clean = re.sub(r"\s+", " ", line).strip()
            key = normalize_text(clean)
            if key and key not in seen and key not in {
                normalize_text(existing) for existing in supplemental_lines
            }:
                supplemental_lines.append(clean)

    return (
        "\n".join(primary_lines),
        best["words"],
        best["lang"],
        "\n".join(supplemental_lines),
    )


def _ocr_image(image):
    text, _words, _lang, _supplemental = _ocr_image_with_data(image)
    return text


def _render_pdf_page(page):
    pixmap = page.get_pixmap(
        matrix=fitz.Matrix(6.0, 6.0),
        alpha=False
    )
    return Image.open(
        io.BytesIO(pixmap.tobytes("png"))
    ).convert("RGB")


def _image_to_png_bytes(image):
    output = BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


def get_output_page_count(file):
    name = str(getattr(file, "name", "")).casefold()

    if name.endswith(".pdf"):
        file.seek(0)
        data = file.read()
        document = fitz.open(stream=data, filetype="pdf")
        count = len(document)
        document.close()
        file.seek(0)
        return count

    if name.endswith((".jpg", ".jpeg", ".png")):
        return 1

    return 0


def extract_output_pages(file):
    """
    Extract artwork pages with OCR as the primary source.

    Every page also stores the rendered full-page image and OCR word boxes so
    the exact artwork can later be displayed with highlight overlays.
    """
    name = str(getattr(file, "name", "")).casefold()

    if name.endswith(".pdf"):
        file.seek(0)
        data = file.read()
        if not data:
            raise ValueError("The Output Artwork PDF is empty.")

        document = fitz.open(stream=data, filetype="pdf")
        pages = []

        try:
            for page_number, page in enumerate(document, start=1):
                direct_text = page.get_text("text") or ""
                image = _render_pdf_page(page)
                ocr_text = ""
                ocr_words = []
                ocr_error = None
                ocr_lang = ""

                try:
                    (
                        ocr_text,
                        ocr_words,
                        ocr_lang,
                        ocr_alt_text,
                    ) = _ocr_image_with_data(image)
                except Exception as exc:
                    ocr_error = exc

                ocr_scale = 1600 / max(1, image.width) if image.width < 1600 else 1.0

                if _usable_text(ocr_text):
                    text = ocr_text
                    source_type = "ocr"
                elif _usable_text(direct_text):
                    text = direct_text
                    source_type = "pdf_text"
                else:
                    detail = str(ocr_error) if ocr_error else "no usable text"
                    raise RuntimeError(
                        f"Page {page_number}: no readable artwork text was found. {detail}"
                    )

                pages.append({
                    "page": page_number,
                    "text": str(text),
                    "source_type": source_type,
                    "direct_text": str(direct_text or ""),
                    "ocr_text": str(ocr_text or ""),
                    "ocr_alt_text": str(ocr_alt_text or ""),
                    "ocr_words": ocr_words,
                    "ocr_lang": ocr_lang,
                    "ocr_scale_x": ocr_scale,
                    "ocr_scale_y": ocr_scale,
                    "image_bytes": _image_to_png_bytes(image),
                    "image_width": image.width,
                    "image_height": image.height,
                })
        finally:
            document.close()

        file.seek(0)
        return pages

    if name.endswith((".jpg", ".jpeg", ".png")):
        file.seek(0)
        image = Image.open(file).convert("RGB")
        (
            ocr_text,
            ocr_words,
            ocr_lang,
            ocr_alt_text,
        ) = _ocr_image_with_data(image)
        if not _usable_text(ocr_text):
            raise RuntimeError("No readable artwork text was detected in the image.")
        ocr_scale = 1600 / max(1, image.width) if image.width < 1600 else 1.0
        file.seek(0)
        return [{
            "page": 1,
            "text": str(ocr_text),
            "source_type": "ocr",
            "direct_text": "",
            "ocr_text": str(ocr_text),
            "ocr_alt_text": str(ocr_alt_text),
            "ocr_words": ocr_words,
            "ocr_lang": ocr_lang,
            "ocr_scale_x": ocr_scale,
            "ocr_scale_y": ocr_scale,
            "image_bytes": _image_to_png_bytes(image),
            "image_width": image.width,
            "image_height": image.height,
        }]

    raise ValueError(
        "Unsupported output format. Please upload PDF, JPG, JPEG, or PNG."
    )


# =========================================================
# VISUAL ARTWORK HIGHLIGHTING
# =========================================================

# Presentation-only palette.  The comparison engine never uses these colors.
# The same field receives the same color in the artwork, table swatch, and
# Excel report.
FIELD_VISUAL_COLORS = [
    "#2563EB",  # blue
    "#0F766E",  # teal
    "#7C3AED",  # violet
    "#DB2777",  # pink
    "#EA580C",  # orange
    "#CA8A04",  # gold
    "#0891B2",  # cyan
    "#C026D3",  # fuchsia
    "#65A30D",  # lime
    "#DC2626",  # red
    "#4F46E5",  # indigo
    "#B45309",  # amber
    "#15803D",  # dark green
    "#0284C7",  # sky blue
    "#BE123C",  # rose
    "#4338CA",  # dark indigo
    "#16A34A",  # green
    "#9D174D",  # deep pink
]


def get_field_visual_colors(fields):
    """Return a stable FIELD -> HEX map based only on displayed field order."""
    colors = {}
    for index, field in enumerate([str(x) for x in (fields or [])]):
        colors[field] = FIELD_VISUAL_COLORS[index % len(FIELD_VISUAL_COLORS)]
    return colors


def add_visual_column(report, selected_fields):
    """Add a presentation-only VISUAL swatch column without changing the report."""
    if report is None:
        return report

    display = report.copy()
    colors = get_field_visual_colors(selected_fields)
    # Streamlit dataframe receives a styling layer for the swatch column.
    # Keep the underlying value empty so Excel can use a true solid-color cell
    # instead of a text glyph that can make every swatch look similar.
    swatches = ["" for _field in display.get("FIELD", []).tolist()]

    if "VISUAL" in display.columns:
        display["VISUAL"] = swatches
    else:
        insert_at = display.columns.get_loc("STATUS") if "STATUS" in display.columns else len(display.columns)
        display.insert(insert_at, "VISUAL", swatches)
    return display


def _visual_norm(text):
    if text is None:
        return ""
    value = unicodedata.normalize("NFKC", str(text)).casefold()
    value = value.replace("%", "")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _visual_compact(text):
    return re.sub(r"[^a-z0-9]", "", _visual_norm(text))


def _page_ocr_scale(page):
    """Return OCR-coordinate -> stored-image coordinate scale factors."""
    try:
        sx = float(page.get("ocr_scale_x", 1.0) or 1.0)
        sy = float(page.get("ocr_scale_y", 1.0) or 1.0)
    except Exception:
        sx = sy = 1.0
    if sx <= 0:
        sx = 1.0
    if sy <= 0:
        sy = 1.0
    return sx, sy


def _scaled_word(word, sx, sy):
    """Convert OCR coordinates back to the original stored artwork image."""
    return {
        **word,
        "left": int(round(float(word.get("left", 0)) / sx)),
        "top": int(round(float(word.get("top", 0)) / sy)),
        "width": max(1, int(round(float(word.get("width", 1)) / sx))),
        "height": max(1, int(round(float(word.get("height", 1)) / sy))),
    }


def _visual_group_words(page):
    """Return OCR words grouped by physical line in stored-image coordinates."""
    sx, sy = _page_ocr_scale(page)
    grouped = {}

    for raw_word in page.get("ocr_words", []) or []:
        if not isinstance(raw_word, dict):
            continue
        if not str(raw_word.get("text", "")).strip():
            continue
        word = _scaled_word(raw_word, sx, sy)
        key = (
            raw_word.get("block_num", 0),
            raw_word.get("par_num", 0),
            raw_word.get("line_num", 0),
        )
        grouped.setdefault(key, []).append(word)

    groups = list(grouped.values())
    for group in groups:
        group.sort(key=lambda item: (float(item.get("left", 0)), float(item.get("top", 0))))

    groups.sort(
        key=lambda group: (
            min(float(word.get("top", 0)) for word in group),
            min(float(word.get("left", 0)) for word in group),
        )
    )
    return groups


def _boxes_from_words(words):
    if not words:
        return []
    left = min(int(word.get("left", 0)) for word in words)
    top = min(int(word.get("top", 0)) for word in words)
    right = max(int(word.get("left", 0)) + int(word.get("width", 0)) for word in words)
    bottom = max(int(word.get("top", 0)) + int(word.get("height", 0)) for word in words)
    return [(left, top, right, bottom)] if right > left and bottom > top else []


def _find_visual_boxes(page, target, field_name=""):
    """
    Presentation-only lookup of the actual OCR words corresponding to the
    comparison result's PDF OUTPUT value.

    IMPORTANT: OCR boxes are converted back from the enlarged OCR image to the
    original stored artwork image before drawing.  This is the key protection
    against the offset caused by the OCR preprocessing resize.
    """
    target_norm = _visual_norm(target)
    target_compact = _visual_compact(target)
    if not target_norm or target_norm in {"not found", "-", "—"}:
        return []

    groups = _visual_group_words(page)
    if not groups:
        return []

    # 1) Exact contiguous OCR-word sequence.
    for group in groups:
        normalized = [_visual_norm(word.get("text", "")) for word in group]
        for start in range(len(group)):
            accumulated = []
            selected = []
            for end in range(start, len(group)):
                token = normalized[end]
                if not token:
                    continue
                accumulated.append(token)
                selected.append(group[end])
                joined = " ".join(accumulated).strip()

                if joined == target_norm or _visual_compact(joined) == target_compact:
                    return _boxes_from_words(selected)

                # Prevent a search from walking through unrelated later words.
                if len(_visual_compact(joined)) > len(target_compact) + 8:
                    break

    # 2) Numeric component inside a combined token such as 44-12.
    if re.fullmatch(r"\d+", target_norm):
        for group in groups:
            for word in group:
                raw = str(word.get("text", "")).strip()
                match = re.search(r"(\d+)[-/](\d+)", raw)
                if not match:
                    continue

                first, second = match.group(1), match.group(2)
                if target_norm not in {first, second}:
                    continue

                left = int(word.get("left", 0))
                top = int(word.get("top", 0))
                width = int(word.get("width", 0))
                height = int(word.get("height", 0))
                full = f"{first}-{second}"
                total_chars = max(1, len(full))

                if target_norm == first:
                    end_x = left + max(1, int(round(width * len(first) / total_chars)))
                    return [(left, top, min(left + width, end_x), top + height)]

                start_x = left + max(1, int(round(width * (len(first) + 1) / total_chars)))
                return [(min(left + width - 1, start_x), top, left + width, top + height)]

    # 3) Controlled cross-line exact compact sequence.
    ordered = [word for group in groups for word in group]
    for start in range(len(ordered)):
        compact = ""
        selected = []
        for end in range(start, min(len(ordered), start + 16)):
            token = _visual_compact(ordered[end].get("text", ""))
            if not token:
                continue
            compact += token
            selected.append(ordered[end])
            if compact == target_compact:
                return _boxes_from_words(selected)
            if len(compact) > len(target_compact) + 8:
                break

    return []


def _hex_rgb(hex_color):
    value = str(hex_color).lstrip("#")
    if len(value) != 6:
        return (37, 99, 235)
    try:
        return tuple(int(value[index:index + 2], 16) for index in (0, 2, 4))
    except Exception:
        return (37, 99, 235)


def _load_visual_fonts():
    try:
        from PIL import ImageFont
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        ]
        bold_candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
        ]
        regular_path = next((p for p in candidates if Path(p).exists()), None)
        bold_path = next((p for p in bold_candidates if Path(p).exists()), None)
        if regular_path and bold_path:
            return (
                ImageFont.truetype(regular_path, 22),
                ImageFont.truetype(bold_path, 22),
            )
    except Exception:
        pass
    return None, None


def _text_size(draw, text, font):
    if font is None:
        return max(20, len(str(text)) * 11), 20
    try:
        bbox = draw.textbbox((0, 0), text, font=font)
        return bbox[2] - bbox[0], bbox[3] - bbox[1]
    except Exception:
        return max(20, len(str(text)) * 11), 20


def _rect_intersects(a, b, pad=8):
    return not (
        a[2] + pad < b[0]
        or a[0] - pad > b[2]
        or a[3] + pad < b[1]
        or a[1] - pad > b[3]
    )


def _place_label_above_or_below(box, label_w, label_h, image_w, image_h, occupied):
    left, top, right, bottom = box
    gap = 10
    x = max(6, min(image_w - label_w - 6, int((left + right - label_w) / 2)))

    candidates = [
        (x, top - label_h - gap),
        (x, bottom + gap),
        (max(6, left - label_w - gap), int((top + bottom - label_h) / 2)),
        (min(image_w - label_w - 6, right + gap), int((top + bottom - label_h) / 2)),
    ]

    for candidate in candidates:
        cx, cy = candidate
        if cx < 6 or cy < 6 or cx + label_w > image_w - 6 or cy + label_h > image_h - 6:
            continue
        rect = (cx, cy, cx + label_w, cy + label_h)
        if not any(_rect_intersects(rect, existing, pad=6) for existing in occupied):
            return candidate

    return (
        max(6, min(image_w - label_w - 6, x)),
        max(6, min(image_h - label_h - 6, top - label_h - gap)),
    )


def _visual_osz_candidates(page):
    """Build visual-only OSZ sequence candidates from scaled OCR coordinates.

    This function is presentation-only. It never participates in PASS/FAIL
    comparison decisions. It exists so an OSZ field can point to the correct
    sequence position even when the comparison result is FAIL.
    """
    numeric_words = []
    for group in _visual_group_words(page):
        for word in group:
            raw = str(word.get("text", "")).strip()
            if not re.fullmatch(r"\d+", raw):
                continue
            width = int(word.get("width", 0) or 0)
            height = int(word.get("height", 0) or 0)
            if width <= 0 or height <= 0:
                continue
            numeric_words.append({
                **word,
                "value": raw,
                "cx": float(word.get("left", 0)) + width / 2.0,
                "cy": float(word.get("top", 0)) + height / 2.0,
            })

    if len(numeric_words) < 3:
        return []

    candidates = []
    for orientation in ("vertical", "horizontal"):
        # A vertical OSZ list progresses on Y while its items share a common X.
        # A horizontal list progresses on X while its items share a common Y.
        axis = "cy" if orientation == "vertical" else "cx"
        cross = "cx" if orientation == "vertical" else "cy"
        ordered = sorted(numeric_words, key=lambda item: (item[cross], item[axis]))

        groups = []
        for word in ordered:
            added = False
            for group in groups:
                ref = sum(item[cross] for item in group) / len(group)
                tolerance = max(
                    14.0,
                    min(90.0, max(word.get("width", 1), word.get("height", 1)) * 1.7)
                )
                if abs(word[cross] - ref) <= tolerance:
                    group.append(word)
                    added = True
                    break
            if not added:
                groups.append([word])

        for group in groups:
            if len(group) < 3:
                continue
            seq = sorted(group, key=lambda item: (item["cy"], item["cx"])) if orientation == "vertical" else sorted(group, key=lambda item: (item["cx"], item["cy"]))
            gaps = [
                (seq[i + 1][axis] - seq[i][axis])
                for i in range(len(seq) - 1)
            ]
            gaps = [gap for gap in gaps if gap > 0]
            if not gaps:
                continue
            median_gap = sorted(gaps)[len(gaps) // 2]
            if median_gap <= 0:
                continue
            regularity = sum(
                0.45 * median_gap <= gap <= 1.85 * median_gap
                for gap in gaps
            ) / len(gaps)
            cross_spread = max(item[cross] for item in seq) - min(item[cross] for item in seq)
            alignment_ratio = cross_spread / max(1.0, median_gap)
            if regularity < 0.55 or alignment_ratio > 0.75:
                continue
            score = len(seq) * 10.0 + regularity * 12.0 - alignment_ratio * 8.0
            signature = tuple((item["value"], round(item["cx"] / 5), round(item["cy"] / 5)) for item in seq)
            if any(existing["signature"] == signature for existing in candidates):
                continue
            candidates.append({
                "orientation": orientation,
                "items": seq,
                "score": score,
                "signature": signature,
            })

    candidates.sort(key=lambda item: item["score"], reverse=True)
    return candidates[:20]


def _find_visual_osz_box(page, field_name, actual_value=""):
    """Find the artwork box for an OSZ field by sequence position."""
    field_type = get_field_type(field_name)
    if field_type != "OSZ":
        return []

    compact = normalize_text(field_name).replace(" ", "")
    match = re.fullmatch(r"osz(\d+)", compact)
    if not match:
        return []
    index = int(match.group(1))
    if index <= 0:
        return []

    candidates = [c for c in _visual_osz_candidates(page) if len(c.get("items", [])) >= index]
    if not candidates:
        return []

    actual_num = normalize_numeric(actual_value)
    if actual_num is not None:
        matching = [
            c for c in candidates
            if normalize_numeric(c["items"][index - 1].get("value")) == actual_num
        ]
        if matching:
            candidates = matching

    chosen = max(candidates, key=lambda item: item.get("score", 0.0))
    item = chosen["items"][index - 1]
    return _boxes_from_words([item])


def _truncate_visual_value(value, limit=42):
    value = re.sub(r"\s+", " ", str(value or "").strip())
    if len(value) <= limit:
        return value
    return value[: max(1, limit - 1)] + "…"


def _visual_field_uses_block_mapping(field_name, expected_value="", actual_value=""):
    """Return True when the visual should show the whole compared text block.

    This is presentation-only. It never changes the comparison decision.
    Long structured text fields are intentionally shown as complete regions so
    the reviewer can see the exact block that was evaluated, rather than only
    one differing word.
    """
    field_type = get_field_type(field_name)
    if field_type in {"CARE", "CONTENT"}:
        return True

    if field_type in {"ATTRIBUTE", "GENERAL", "BRAND"}:
        sample = str(actual_value or expected_value or "").strip()
        tokens = tokenize(sample)
        return len(tokens) >= 5 or len(sample) >= 45

    return False


def _visual_group_text(group):
    return re.sub(
        r"\s+",
        " ",
        " ".join(str(word.get("text", "")).strip() for word in group if str(word.get("text", "")).strip())
    ).strip()


def _visual_match_line_score(target_line, actual_line):
    """Similarity score for locating one already-compared OCR line."""
    target_norm = _visual_norm(target_line)
    actual_norm = _visual_norm(actual_line)
    if not target_norm or not actual_norm:
        return 0.0

    if target_norm == actual_norm or _visual_compact(target_norm) == _visual_compact(actual_norm):
        return 1.0

    from difflib import SequenceMatcher

    target_tokens = target_norm.split()
    actual_tokens = actual_norm.split()
    token_ratio = SequenceMatcher(
        None, target_tokens, actual_tokens, autojunk=False
    ).ratio()
    compact_ratio = SequenceMatcher(
        None, _visual_compact(target_norm), _visual_compact(actual_norm), autojunk=False
    ).ratio()

    # Token similarity is more meaningful for long care/composition lines.
    return max(token_ratio, compact_ratio * 0.98)


def _visual_region_markers(field_name):
    """Return start/stop markers for locating a complete visual text region."""
    field_type = get_field_type(field_name)
    region = get_field_region(field_name)

    if field_type == "CARE":
        starts = {
            "EN": ["machine wash", "wash", "bleach", "dry clean", "tumble dry", "cool iron"],
            "FR": ["laver", "blanchiment", "nettoyage", "sécher", "repasser"],
            "SP": ["lavar", "cloro", "secadora", "plancha", "limpieza en seco"],
            "": ["machine wash", "wash", "bleach", "dry clean", "laver", "lavar"],
        }
        return starts.get(region, starts[""]), [
            "rn", "rn#", "made in", "fabrique en", "hecho en"
        ]

    if field_type == "CONTENT":
        return ["%", "cotton", "polyester", "spandex", "elastane", "nylon", "shell", "lining"], [
            "rn", "rn#", "made in", "fabrique en", "hecho en",
            "machine wash", "wash", "laver", "lavar", "bleach", "cloro"
        ]

    return [], ["rn", "rn#", "made in", "fabrique en", "hecho en"]


def _visual_is_numeric_only_group(group):
    text = _visual_group_text(group)
    compact = re.sub(r"[^0-9]", "", text)
    alpha = re.sub(r"[^a-z]", "", text.casefold())
    return bool(compact) and not alpha and len(compact) <= 6


def _visual_is_short_code_group(group):
    """Detect small standalone technical codes after a long text block."""
    text = _visual_group_text(group).strip()
    compact_alpha = re.sub(r"[^A-Za-z]", "", text)
    if not compact_alpha:
        return False
    if len(compact_alpha) > 5:
        return False
    # Preserve normal short words that legitimately continue care text.
    common_short_words = {
        "if", "no", "use", "with", "cold", "like", "only", "when",
        "dry", "low", "iron", "cool", "wash", "bleach", "needed",
        "laver", "avec", "sans", "si", "lavar"
    }
    normalized = _visual_norm(text)
    if normalized in common_short_words:
        return False
    return text.upper() == text and len(compact_alpha) <= 5


def _find_visual_semantic_block_boxes(page, field_name):
    """Locate the complete physical artwork block for a long text field.

    This is presentation-only. It never changes PASS/FAIL logic. For CARE the
    physical block can be interrupted by standalone OSZ numbers, so numeric-only
    lines are skipped while the care text continues. For CONTENT, the block ends
    when the next care/COO/RN region starts.
    """
    groups = _visual_group_words(page)
    if not groups:
        return []

    starts, stops = _visual_region_markers(field_name)
    starts = [_visual_norm(x) for x in starts if _visual_norm(x)]
    stops = [_visual_norm(x) for x in stops if _visual_norm(x)]
    if not starts:
        return []

    field_type = get_field_type(field_name)
    start_idx = None
    for idx, group in enumerate(groups):
        line_text = _visual_norm(_visual_group_text(group))
        if line_text and any(marker in line_text for marker in starts):
            start_idx = idx
            break

    if start_idx is None:
        return []

    selected_groups = []
    meaningful_text_groups = 0

    for idx in range(start_idx, len(groups)):
        group = groups[idx]
        line_text = _visual_norm(_visual_group_text(group))

        if idx > start_idx and any(marker in line_text for marker in stops):
            break

        # A numeric OSZ line may sit inside/adjacent to the care block. It is
        # not part of the care instruction and must not be highlighted.
        if idx > start_idx and _visual_is_numeric_only_group(group):
            continue

        if idx > start_idx and field_type == "CARE" and _visual_is_short_code_group(group):
            break

        if line_text:
            selected_groups.append(group)
            meaningful_text_groups += 1

        if meaningful_text_groups >= 20:
            break

    boxes = []
    for group in selected_groups:
        box = _boxes_from_words(group)
        if box:
            boxes.extend(box)
    return boxes


def _find_visual_block_boxes(page, target, field_name=""):
    """Locate the complete OCR line block represented by a compared text value.

    For CARE/CONTENT (and long ATTRIBUTE/GENERAL/BRAND fields), the comparison
    result often represents several complete artwork lines. The visual layer
    therefore maps every relevant OCR line in that block, rather than searching
    only for the word that differs.
    """
    target = str(target or "").strip()
    if not target or target.casefold() in {"not found", "-", "—"}:
        return []

    groups = _visual_group_words(page)
    if not groups:
        return []

    target_lines = [
        re.sub(r"\s+", " ", line).strip()
        for line in target.splitlines()
        if str(line).strip()
    ]

    # When the comparison PDF OUTPUT is a single long line, wrap it only for
    # matching if the OCR itself has multiple physical lines. Otherwise use the
    # whole matched OCR group.
    if not target_lines:
        return []

    group_texts = [_visual_group_text(group) for group in groups]

    # First pass: exact/compact line matches, preserving physical order.
    chosen_indices = []
    search_from = 0
    for target_line in target_lines:
        exact_idx = None
        for idx in range(search_from, len(groups)):
            candidate = group_texts[idx]
            if not candidate:
                continue
            if (
                _visual_norm(target_line) == _visual_norm(candidate)
                or _visual_compact(target_line) == _visual_compact(candidate)
            ):
                exact_idx = idx
                break
        if exact_idx is not None:
            chosen_indices.append(exact_idx)
            search_from = exact_idx + 1

    # If all target lines were found, return those physical lines.
    if len(chosen_indices) == len(target_lines):
        selected_groups = [groups[idx] for idx in chosen_indices]
        boxes = []
        for group in selected_groups:
            boxes.extend(_boxes_from_words(group))
        return boxes

    # Second pass: fuzzy line alignment. Use a monotonic assignment so the
    # highlight cannot jump to a distant unrelated block just because one word
    # happens to be similar.
    fuzzy_indices = []
    search_from = 0
    for target_line in target_lines:
        best_idx = None
        best_score = 0.0
        for idx in range(search_from, min(len(groups), search_from + 12)):
            score = _visual_match_line_score(target_line, group_texts[idx])
            if score > best_score:
                best_score = score
                best_idx = idx
        if best_idx is None or best_score < 0.68:
            continue
        fuzzy_indices.append(best_idx)
        search_from = best_idx + 1

    if not fuzzy_indices:
        return []

    # For a block field, at least half the source lines should align before we
    # accept the result. This protects against accidental one-word matches.
    if len(fuzzy_indices) < max(1, int(len(target_lines) * 0.5)):
        return []

    # Include intermediate physical OCR lines between first and last aligned
    # targets when the comparison target represents a continuous block. This is
    # what makes a 5-line care instruction appear as one clearly compared region.
    first_idx = fuzzy_indices[0]
    last_idx = fuzzy_indices[-1]
    selected_groups = groups[first_idx:last_idx + 1]

    boxes = []
    for group in selected_groups:
        box = _boxes_from_words(group)
        if box:
            boxes.extend(box)
    return boxes


def _find_visual_failure_boxes(page, field_name, expected_value, actual_value, difference):
    """Find the actual artwork region for a FAIL, presentation layer only."""
    actual_value = str(actual_value or "").strip()
    expected_value = str(expected_value or "").strip()
    difference = str(difference or "").strip()

    if get_field_type(field_name) == "OSZ":
        boxes = _find_visual_osz_box(page, field_name, actual_value)
        if boxes:
            return boxes

    # For block-based fields, ALWAYS try the complete compared block first.
    # This intentionally prevents a care mismatch from highlighting only a word
    # such as "ONLY".
    if _visual_field_uses_block_mapping(field_name, expected_value, actual_value):
        boxes = _find_visual_semantic_block_boxes(page, field_name)
        if not boxes:
            boxes = _find_visual_block_boxes(page, actual_value, field_name)
        if boxes:
            return boxes

    # Fallback for scalar/structured failures: locate the actual PDF output.
    boxes = _find_visual_boxes(page, actual_value, field_name)
    if boxes:
        return boxes

    from difflib import SequenceMatcher

    expected_tokens = tokenize(expected_value)
    actual_tokens = tokenize(actual_value)
    candidate_tokens = []

    if expected_tokens and actual_tokens:
        matcher = SequenceMatcher(None, expected_tokens, actual_tokens, autojunk=False)
        for tag, _a1, _a2, b1, b2 in matcher.get_opcodes():
            if tag in {"replace", "insert"}:
                candidate_tokens.extend(actual_tokens[b1:b2])

    # Stored differences can contain useful actual-value fragments.
    for item in re.findall(
        r"(?:Extra|Found|PDF)\s*:\s*([^;]+)",
        difference,
        flags=re.IGNORECASE
    ):
        candidate_tokens.extend(tokenize(item))

    seen = set()
    cleaned = []
    for token in candidate_tokens:
        token_norm = normalize_text(token)
        if not token_norm:
            continue
        if len(token_norm) <= 1 and len(actual_tokens) > 1:
            continue
        if token_norm not in seen:
            seen.add(token_norm)
            cleaned.append(token_norm)

    found = []
    for token in cleaned:
        for box in _find_visual_boxes(page, token, field_name):
            if box not in found:
                found.append(box)

    return found


def build_highlighted_page_image(page, page_report, field_colors=None):
    """Render the original artwork with precise field mapping annotations.

    Presentation-only layer. Comparison decisions are read from ``page_report``
    and are never recalculated or modified here.
    """
    image_bytes = page.get("image_bytes")
    if not image_bytes:
        return None

    base = Image.open(BytesIO(image_bytes)).convert("RGBA")
    overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    field_colors = field_colors or {}
    regular_font, bold_font = _load_visual_fonts()

    if page_report is None or page_report.empty:
        return base.convert("RGB")

    annotations = []
    field_order = []
    for _, row in page_report.iterrows():
        field = str(row.get("FIELD", ""))
        if field and field not in field_order:
            field_order.append(field)

    number_by_field = {field: index + 1 for index, field in enumerate(field_order)}
    annotated_fields = set()

    for _, row in page_report.iterrows():
        status = str(row.get("STATUS", "")).strip()
        if status not in {"PASS", "FAIL"}:
            continue

        field_name = str(row.get("FIELD", "")).strip()
        pdf_output = str(row.get("PDF OUTPUT", "")).strip()
        expected_value = str(row.get("ORDER FORM DATA", "")).strip()

        if not pdf_output or pdf_output.casefold() in {"not found", "—", "-"}:
            continue

        # Block fields are mapped to the whole compared physical region.
        # Prefer semantic artwork-region detection because PDF OUTPUT may be a
        # comparison-constructed long string rather than a verbatim OCR line.
        if _visual_field_uses_block_mapping(field_name, expected_value, pdf_output):
            boxes = _find_visual_semantic_block_boxes(page, field_name)
            if not boxes:
                boxes = _find_visual_block_boxes(page, pdf_output, field_name)
        elif get_field_type(field_name) == "OSZ":
            boxes = _find_visual_osz_box(page, field_name, pdf_output)
        else:
            boxes = _find_visual_boxes(page, pdf_output, field_name)

        if status == "FAIL" and not boxes:
            boxes = _find_visual_failure_boxes(
                page,
                field_name,
                expected_value,
                pdf_output,
                row.get("DIFFERENCE", ""),
            )

        if not boxes:
            continue

        color = field_colors.get(field_name, FIELD_VISUAL_COLORS[0])
        rgb = _hex_rgb(color)

        for box in boxes:
            left, top, right, bottom = box
            pad = max(2, int(min(base.size) * 0.0018))
            left = max(0, left - pad)
            top = max(0, top - pad)
            right = min(base.width - 1, right + pad)
            bottom = min(base.height - 1, bottom + pad)

            if status == "FAIL":
                # Red is the defect color; the field-specific outline preserves
                # field identity.
                draw.rounded_rectangle(
                    (left, top, right, bottom),
                    radius=max(3, pad),
                    fill=(218, 54, 51, 100),
                    outline=rgb + (255,),
                    width=max(2, pad),
                )
                draw.rounded_rectangle(
                    (max(0, left - 1), max(0, top - 1),
                     min(base.width - 1, right + 1), min(base.height - 1, bottom + 1)),
                    radius=max(3, pad + 1),
                    outline=(218, 54, 51, 255),
                    width=max(2, pad // 2),
                )
            else:
                draw.rounded_rectangle(
                    (left, top, right, bottom),
                    radius=max(3, pad),
                    fill=rgb + (96,),
                    outline=rgb + (255,),
                    width=max(2, pad),
                )

            if field_name not in annotated_fields:
                annotations.append({
                    "field": field_name,
                    "status": status,
                    "color": color,
                    "rgb": rgb,
                    "box": (left, top, right, bottom),
                    "number": number_by_field.get(field_name, 0),
                })
                annotated_fields.add(field_name)

    # Labels intentionally remain simple: the artwork communicates the scope
    # of comparison, while detailed reasons stay in the comparison table.
    used_label_rects = []
    for item in annotations:
        box = item["box"]
        field_name = item["field"]
        status = item["status"]
        rgb = item["rgb"]
        number = item["number"]

        title_text = f"{number}. {field_name} • {status}"
        font = bold_font or regular_font
        tw, th = _text_size(draw, title_text, font)

        max_label_w = max(210, int(base.width * 0.38))
        label_w = min(tw + 26, max_label_w)
        label_h = th + 16

        label_x, label_y = _place_label_above_or_below(
            box,
            label_w,
            label_h,
            base.width,
            base.height,
            used_label_rects,
        )
        used_label_rects.append((label_x, label_y, label_x + label_w, label_y + label_h))

        box_cx = int((box[0] + box[2]) / 2)
        if label_y + label_h <= box[1]:
            label_anchor = (int(label_x + label_w / 2), int(label_y + label_h))
            target_anchor = (box_cx, box[1])
        elif label_y >= box[3]:
            label_anchor = (int(label_x + label_w / 2), int(label_y))
            target_anchor = (box_cx, box[3])
        elif label_x + label_w <= box[0]:
            label_anchor = (int(label_x + label_w), int(label_y + label_h / 2))
            target_anchor = (box[0], int((box[1] + box[3]) / 2))
        else:
            label_anchor = (int(label_x), int(label_y + label_h / 2))
            target_anchor = (box[2], int((box[1] + box[3]) / 2))

        draw.line(
            (label_anchor[0], label_anchor[1], target_anchor[0], target_anchor[1]),
            fill=((218, 54, 51) if status == "FAIL" else rgb) + (235,),
            width=max(2, int(min(base.size) * 0.00095)),
        )

        label_outline = (218, 54, 51) if status == "FAIL" else rgb
        draw.rounded_rectangle(
            (label_x, label_y, label_x + label_w, label_y + label_h),
            radius=max(5, int(min(base.size) * 0.0025)),
            fill=(255, 255, 255, 246),
            outline=label_outline + (255,),
            width=max(2, int(min(base.size) * 0.0012)),
        )

        bar_w = max(5, int(label_w * 0.02))
        draw.rounded_rectangle(
            (label_x, label_y, label_x + bar_w, label_y + label_h),
            radius=max(2, int(bar_w * 0.35)),
            fill=rgb + (255,),
        )

        draw.text(
            (label_x + bar_w + 8, label_y + 6),
            title_text,
            fill=(20, 28, 40, 255),
            font=font,
        )

    return Image.alpha_composite(base, overlay).convert("RGB")


def _visual_image_bytes(image):
    if image is None:
        return None
    return _image_to_png_bytes(image)

def clean_pdf_line(line):
    if not line:
        return ""
    line = str(line).replace("\u200b", "").replace("\ufeff", "").strip()
    line = re.sub(r"^\s*n\s+(?=[A-Za-z])", "", line)
    return line.strip()


# =========================================================
# ATOMIC PAGE MODEL
# =========================================================

def build_page_lines(page_text, product_type):
    """
    Create atomic lines. We do NOT create overlapping blocks and then consume
    blocks, because consuming an overlapping block can accidentally consume
    unrelated values. Instead, every decision is tied to actual line IDs.
    """
    raw_lines = str(page_text or "").splitlines()
    lines = []

    panel_pattern = re.compile(
        r"^\s*(?:panel\s*)?(\d{1,3})\s*$",
        re.IGNORECASE
    )

    for raw_index, raw in enumerate(raw_lines):
        line = clean_pdf_line(raw)
        if not line:
            continue
        if len(line) > 1500:
            continue

        # PFL: remove only explicit panel labels. Do not discard standalone
        # numeric lines because those may be genuine OSZ/size data.
        if product_type == "PFL" and re.match(r"^\s*panel\s*[-#: ]?\d{1,3}\s*$", line, re.IGNORECASE):
            continue

        lines.append({
            "line_id": len(lines),
            "raw_index": raw_index,
            "text": line,
            "norm": normalize_text(line)
        })

    return lines


def build_page_state(page, product_type):
    if not isinstance(page, dict):
        raise TypeError(
            f"Output page data is malformed. Expected a page dictionary, got {type(page).__name__}."
        )

    lines = build_page_lines(
        page.get("text", ""),
        product_type
    )

    # Defensive validation: every line must be a dictionary with the fields
    # consumed by the matching engine. This prevents the vague
    # "list indices must be integers or slices, not str" failure.
    invalid_lines = [
        index for index, line in enumerate(lines)
        if not isinstance(line, dict)
        or "line_id" not in line
        or "text" not in line
        or "norm" not in line
    ]
    if invalid_lines:
        raise TypeError(
            f"Output page {page.get('page', '?')} contains malformed OCR/text lines: {invalid_lines[:5]}"
        )

    return {
        "page": page.get("page"),
        "source_type": page.get("source_type", "pdf_text"),
        "lines": lines,
        "consumed": set(),
        "consumed_spans": {},
        # Keep the existing line-based OSZ runs intact.  The additional
        # coordinate-aware candidates are used only as a fallback when OCR
        # formatting has mixed multiple standalone numbers into one line.
        "osz_runs": extract_standalone_numeric_runs_from_lines(lines),
        # Keep a second, immutable numeric sequence from the PDF text layer.
        # This is a targeted OSZ fallback only; it does not alter normal field
        # matching or consumption. It protects clean PDF text when OCR has
        # merged some sequence values into neighbouring lines.
        "osz_direct_runs": extract_standalone_numeric_runs_from_lines(
            build_page_lines(page.get("direct_text", ""), product_type)
        ),
        "osz_sequence_candidates": extract_osz_sequence_candidates(page),
    }


def extract_standalone_numeric_runs_from_lines(lines):
    runs = []
    current = []

    for line in lines:
        if re.fullmatch(r"[-+]?\d+", line["norm"]):
            current.append(line)
        else:
            if current:
                runs.append(current)
                current = []

    if current:
        runs.append(current)

    return [run for run in runs if run]


def _osz_numeric_words(page):
    """Return standalone OCR integer tokens with their physical coordinates."""
    result = []
    for word in page.get("ocr_words", []) or []:
        if not isinstance(word, dict):
            continue
        raw = str(word.get("text", "")).strip()
        if not re.fullmatch(r"\d+", raw):
            continue
        # Ignore implausibly tiny OCR fragments.
        if int(word.get("width", 0) or 0) <= 0 or int(word.get("height", 0) or 0) <= 0:
            continue
        result.append({
            "value": raw,
            "left": int(word.get("left", 0) or 0),
            "top": int(word.get("top", 0) or 0),
            "width": int(word.get("width", 0) or 0),
            "height": int(word.get("height", 0) or 0),
            "center_x": float(word.get("left", 0) or 0) + float(word.get("width", 0) or 0) / 2.0,
            "center_y": float(word.get("top", 0) or 0) + float(word.get("height", 0) or 0) / 2.0,
        })
    return result


def _score_osz_sequence(items, orientation):
    if len(items) < 3:
        return -1.0

    if orientation == "vertical":
        ordered = sorted(items, key=lambda x: (x["center_y"], x["center_x"]))
        axis = [item["center_y"] for item in ordered]
        cross = [item["center_x"] for item in ordered]
    else:
        ordered = sorted(items, key=lambda x: (x["center_x"], x["center_y"]))
        axis = [item["center_x"] for item in ordered]
        cross = [item["center_y"] for item in ordered]

    gaps = [axis[i + 1] - axis[i] for i in range(len(axis) - 1)]
    positive_gaps = [g for g in gaps if g > 0]
    if not positive_gaps:
        return -1.0

    median_gap = sorted(positive_gaps)[len(positive_gaps) // 2]
    if median_gap <= 0:
        return -1.0

    cross_spread = max(cross) - min(cross)
    # OSZ lists can have some OCR jitter, but should remain visually aligned.
    alignment_penalty = cross_spread / max(1.0, median_gap)

    regularity = sum(
        1.0
        for gap in positive_gaps
        if 0.45 * median_gap <= gap <= 1.80 * median_gap
    ) / len(positive_gaps)

    # Prefer longer, well-aligned runs.  A minimum regularity prevents random
    # page numbers or price fragments from becoming an OSZ sequence.
    if regularity < 0.60 or alignment_penalty > 0.55:
        return -1.0

    return len(ordered) * 10.0 + regularity * 10.0 - alignment_penalty * 5.0


def extract_osz_sequence_candidates(page):
    """Build dynamic OSZ candidates from artwork geometry, independent of field count."""
    words = _osz_numeric_words(page)
    if len(words) < 3:
        return []

    candidates = []
    for orientation in ("vertical", "horizontal"):
        # Build loose spatial groups around a common axis.
        if orientation == "vertical":
            words_sorted = sorted(words, key=lambda x: x["center_x"])
            axis_key = "center_x"
        else:
            words_sorted = sorted(words, key=lambda x: x["center_y"])
            axis_key = "center_y"

        groups = []
        for word in words_sorted:
            added = False
            for group in groups:
                reference = sum(item[axis_key] for item in group) / len(group)
                tolerance = max(18.0, min(90.0, max(word["width"], word["height"]) * 1.6))
                if abs(word[axis_key] - reference) <= tolerance:
                    group.append(word)
                    added = True
                    break
            if not added:
                groups.append([word])

        for group in groups:
            score = _score_osz_sequence(group, orientation)
            if score < 0:
                continue
            ordered = (
                sorted(group, key=lambda x: (x["center_y"], x["center_x"]))
                if orientation == "vertical"
                else sorted(group, key=lambda x: (x["center_x"], x["center_y"]))
            )
            # Deduplicate identical sequences.
            signature = tuple((item["value"], round(item["center_x"] / 5), round(item["center_y"] / 5)) for item in ordered)
            if any(existing["signature"] == signature for existing in candidates):
                continue
            candidates.append({
                "orientation": orientation,
                "items": ordered,
                "score": score,
                "signature": signature,
            })

    candidates.sort(key=lambda item: item["score"], reverse=True)
    return candidates[:12]


def _osz_candidate_matches_expected(candidate, index, expected_num):
    items = candidate.get("items", [])
    if len(items) < index:
        return False
    return normalize_numeric(items[index - 1].get("value")) == expected_num


def line_is_available(line, state):
    # A line remains available when only part of it has already been assigned.
    # This is required for combined output lines containing multiple fields.
    return line["line_id"] not in state.get("consumed", set())


def consume_lines(state, lines):
    for line in lines:
        state.setdefault("consumed", set()).add(line["line_id"])


def consume_match(state, match_info):
    """Safely consume either a multi-line match or one matched span."""
    if not isinstance(match_info, dict):
        return

    matched_lines = match_info.get("lines")
    if matched_lines:
        if isinstance(matched_lines, dict):
            matched_lines = [matched_lines]
        valid_lines = [
            line for line in matched_lines
            if isinstance(line, dict) and "line_id" in line
        ]
        if valid_lines:
            consume_lines(state, valid_lines)
            return

    line = match_info.get("line")
    start = match_info.get("start")
    end = match_info.get("end")
    if isinstance(line, dict) and start is not None and end is not None:
        consume_span(state, line, int(start), int(end))


def join_lines(lines):
    return " ".join(
        line["text"]
        for line in lines
    ).strip()


# =========================================================
# SAFE EXACT MATCHING
# =========================================================

def _used_spans(state, line_id):
    return state.setdefault("consumed_spans", {}).get(line_id, [])


def _span_overlaps(a_start, a_end, used_spans):
    return any(a_start < end and a_end > start for start, end in used_spans)


def consume_span(state, line, start, end):
    state.setdefault("consumed_spans", {}).setdefault(line["line_id"], []).append((start, end))


def _normalized_match_positions(expected_norm, actual_norm, field_type):
    if not expected_norm or not actual_norm:
        return []

    # Structured scalar/numeric values must match as whole tokens.
    if normalize_numeric(expected_norm) is not None:
        return [
            (m.start(), m.end())
            for m in re.finditer(
                rf"(?<![A-Za-z0-9]){re.escape(expected_norm)}(?![A-Za-z0-9])",
                actual_norm
            )
        ]

    # Identifiers: normal whole-token match. The asymmetric prefix rule is
    # handled separately by find_identifier_match().
    if field_type == "IDENTIFIER":
        return [
            (m.start(), m.end())
            for m in re.finditer(
                rf"(?<![A-Za-z0-9]){re.escape(expected_norm)}(?![A-Za-z0-9])",
                actual_norm
            )
        ]

    # A single alphanumeric token must not match inside another word.
    if len(expected_norm.split()) == 1 and re.fullmatch(r"[A-Za-z0-9%#]+", expected_norm):
        return [
            (m.start(), m.end())
            for m in re.finditer(
                rf"(?<![A-Za-z0-9]){re.escape(expected_norm)}(?![A-Za-z0-9])",
                actual_norm
            )
        ]

    # Multi-word text can occur inside a combined artwork line.
    return [
        (m.start(), m.end())
        for m in re.finditer(
            re.escape(expected_norm),
            actual_norm
        )
    ]


def safe_exact_match(expected, actual, field_name):
    field_type = get_field_type(field_name)

    if field_type == "SYMBOL":
        exp = normalize_symbol_text(expected)
        act = normalize_symbol_text(actual)
        return bool(exp and exp == act)

    expected_numeric = normalize_numeric(expected)
    if expected_numeric is not None:
        actual_numbers = re.findall(
            r"(?<![A-Za-z0-9])[-+]?\d+(?:\.\d+)?(?![A-Za-z0-9])",
            normalize_text(actual)
        )
        return expected_numeric in {
            normalize_numeric(number)
            for number in actual_numbers
        }

    exp = normalize_text(expected)
    act = normalize_text(actual)

    if not exp or not act:
        return False

    if exp == act:
        return True

    positions = _normalized_match_positions(exp, act, field_type)
    return bool(positions)


def find_identifier_match(expected, state):
    expected_norm = normalize_text(expected)
    if not expected_norm:
        return None

    token_pattern = re.compile(r"[A-Za-z0-9][A-Za-z0-9_\-/]*")

    # First: exact identifier token.
    for line in state["lines"]:
        if not line_is_available(line, state):
            continue

        actual_norm = line["norm"]

        for match in token_pattern.finditer(actual_norm):
            token = match.group(0)
            if token.casefold() != expected_norm.casefold():
                continue

            if _span_overlaps(match.start(), match.end(), _used_spans(state, line["line_id"])):
                continue

            return {
                "kind": "PASS",
                "line": line,
                "start": match.start(),
                "end": match.end(),
                "actual": token,
                "difference": "—",
                "match_type": "IDENTIFIER_EXACT"
            }

    # Second: the Order Form has the base code and the PDF has an additional
    # suffix/static portion. Base -> longer output is intentionally PASS.
    for line in state["lines"]:
        if not line_is_available(line, state):
            continue

        actual_norm = line["norm"]

        for match in token_pattern.finditer(actual_norm):
            token = match.group(0)

            if not token.casefold().startswith(expected_norm.casefold()):
                continue

            if len(token) <= len(expected_norm):
                continue

            if _span_overlaps(match.start(), match.end(), _used_spans(state, line["line_id"])):
                continue

            return {
                "kind": "PASS",
                "line": line,
                "start": match.start(),
                "end": match.start() + len(expected_norm),
                "actual": token,
                "difference": "Base identifier matched; PDF contains additional suffix/static characters.",
                "match_type": "IDENTIFIER_BASE_PLUS_SUFFIX"
            }

    # Third: when both sides are extended identifiers that share a meaningful
    # code prefix but differ, report a genuine FAIL. Example:
    # Order Form USX690 vs PDF USX609.
    for line in state["lines"]:
        if not line_is_available(line, state):
            continue

        actual_norm = line["norm"]
        for match in token_pattern.finditer(actual_norm):
            token = match.group(0)
            if token.casefold() == expected_norm.casefold():
                continue
            common = 0
            for left, right in zip(expected_norm.casefold(), token.casefold()):
                if left != right:
                    break
                common += 1
            if common < 3:
                continue
            if len(token) < 2:
                continue
            if _span_overlaps(match.start(), match.end(), _used_spans(state, line["line_id"])):
                continue

            return {
                "kind": "FAIL",
                "line": line,
                "start": match.start(),
                "end": match.end(),
                "actual": token,
                "difference": (
                    f"Identifier mismatch: Order Form has '{expected}', "
                    f"but PDF has '{token}'."
                ),
                "match_type": "IDENTIFIER_PREFIX_MISMATCH"
            }

    # Fourth: reverse condition — PDF has only a shorter base and Order Form
    # expects the longer identifier. This must NOT pass.
    for line in state["lines"]:
        if not line_is_available(line, state):
            continue

        actual_norm = line["norm"]

        for match in token_pattern.finditer(actual_norm):
            token = match.group(0)

            if len(token) < 2:
                continue

            if not expected_norm.casefold().startswith(token.casefold()):
                continue

            if token.casefold() == expected_norm.casefold():
                continue

            if _span_overlaps(match.start(), match.end(), _used_spans(state, line["line_id"])):
                continue

            return {
                "kind": "FAIL",
                "line": line,
                "start": match.start(),
                "end": match.end(),
                "actual": token,
                "difference": (
                    f"Identifier incomplete: Order Form has '{expected}', "
                    f"but PDF has '{token}'."
                ),
                "match_type": "IDENTIFIER_SHORTER_OUTPUT"
            }

    return None


def find_exact_lines(expected, field_name, state, max_window=8):
    """Find an exact value while allowing multiple independent fields on one PDF line."""
    lines = state["lines"]
    available = [line for line in lines if line_is_available(line, state)]
    if not available:
        return None

    field_type = get_field_type(field_name)

    if field_type == "IDENTIFIER":
        identifier = find_identifier_match(expected, state)
        if identifier and identifier["kind"] == "PASS":
            return [identifier["line"]], identifier
        return None

    numeric_expected = normalize_numeric(expected)
    if numeric_expected is not None:
        for line in available:
            positions = _normalized_match_positions(
                numeric_expected,
                line["norm"],
                field_type
            )
            for start, end in positions:
                if not _span_overlaps(start, end, _used_spans(state, line["line_id"])):
                    return [line], {
                        "kind": "PASS",
                        "line": line,
                        "start": start,
                        "end": end,
                        "actual": numeric_expected,
                        "difference": "—",
                        "match_type": "NUMERIC_TOKEN"
                    }
        return None

    preferred_single_line = field_type in {
        "SYMBOL", "RN", "OSZ", "SIZE", "COLOR", "GENDER",
        "BATCH", "QUANTITY", "BRAND", "ATTRIBUTE", "GENERAL"
    }
    search_window = 1 if preferred_single_line else max_window

    # First search individual lines. This is essential for combined output such
    # as 'SF8334 67 YZP': each value can be consumed independently.
    for line in available:
        positions = _normalized_match_positions(
            normalize_text(expected),
            line["norm"],
            field_type
        )
        for start, end in positions:
            if not _span_overlaps(start, end, _used_spans(state, line["line_id"])):
                return [line], {
                    "kind": "PASS",
                    "line": line,
                    "start": start,
                    "end": end,
                    "actual": line["norm"][start:end],
                    "difference": "—",
                    "match_type": "EXACT_COMBINED_LINE"
                }

    if search_window <= 1:
        return None

    # Multi-line exact match for wrapped care/content/general text.
    for window_size in range(2, search_window + 1):
        for start_index in range(0, len(available) - window_size + 1):
            candidate = available[start_index:start_index + window_size]
            ids = [line["line_id"] for line in candidate]
            if ids != list(range(ids[0], ids[-1] + 1)):
                continue

            text = join_lines(candidate)
            if safe_exact_match(expected, text, field_name):
                return candidate, {
                    "kind": "PASS",
                    "lines": candidate,
                    "difference": "—",
                    "match_type": "EXACT_MULTI_LINE"
                }

    return None

# =========================================================
# STRUCTURED EXTRACTORS
# =========================================================

def extract_coo_value(text):
    normalized = normalize_text(text)
    if not normalized:
        return None

    patterns = [
        r"\bmade\s+in\s+([a-z][a-z\s\-]*)",
        r"\bfabrique\s+en\s+([a-z][a-z\s\-]*)",
        r"\bhecho\s+en\s+([a-z][a-z\s\-]*)",
    ]

    for pattern in patterns:
        match = re.search(pattern, normalized)
        if not match:
            continue
        full = match.group(0).strip()
        full = re.split(
            r"\b(?:rn|ca|sku|size|color|colour|wash|machine)\b",
            full,
            maxsplit=1
        )[0].strip()
        return full

    return None


def coo_language(text):
    normalized = normalize_text(text)
    if "made in" in normalized:
        return "EN"
    if "fabrique en" in normalized:
        return "FR"
    if "hecho en" in normalized:
        return "SP"
    return ""


def extract_rn_value(text):
    normalized = normalize_text(text)
    if not normalized:
        return None

    match = re.search(
        r"\b(?:rn|ca)\s*[#:.-]?\s*([0-9][0-9a-z\-\/]*)",
        normalized,
        re.IGNORECASE
    )
    return match.group(1) if match else None


def extract_identifier_value(text):
    normalized = normalize_text(text)
    patterns = [
        r"\bsku\s*[:#-]?\s*([a-z0-9][a-z0-9_\-/]*)",
        r"\bitem\s*(?:code|no|number)\s*[:#-]?\s*([a-z0-9][a-z0-9_\-/]*)",
        r"\bstyle\s*(?:code|no|number)?\s*[:#-]?\s*([a-z0-9][a-z0-9_\-/]*)",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized, re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def extract_size_value(text):
    normalized = normalize_text(text)
    match = re.search(
        r"\bsize\s*[:#-]?\s*([a-z0-9][a-z0-9\s\-/]*)",
        normalized,
        re.IGNORECASE
    )
    if not match:
        return None
    value = re.split(
        r"\b(?:rn|ca|made|color|colour|sku|style)\b",
        match.group(1),
        maxsplit=1
    )[0].strip()
    return value or None


def extract_color_value(text):
    normalized = normalize_text(text)
    match = re.search(
        r"\b(?:color|colour)\s*[:#-]?\s*([a-z][a-z\s\-/]*)",
        normalized,
        re.IGNORECASE
    )
    if not match:
        return None
    value = re.split(
        r"\b(?:size|rn|ca|made|country|sku|style)\b",
        match.group(1),
        maxsplit=1
    )[0].strip()
    return value or None


def extract_gender_value(text):
    normalized = normalize_text(text)
    for value in [
        "boys", "girls", "women", "men", "unisex",
        "boy", "girl", "woman", "man"
    ]:
        if re.search(rf"\b{re.escape(value)}\b", normalized):
            return value
    return None


def extract_content_values(text):
    normalized = normalize_text(text)
    if not normalized:
        return []

    # Support both normal PDF text ('60% cotton') and OCR/PDF extraction that
    # collapses the percent sign ('60cotton'). Material is captured until the
    # next percentage/material component or end of the content run.
    pattern = re.compile(
        r"(?P<pct>\d{1,3}(?:\.\d+)?)\s*%?\s*"
        r"(?P<material>[a-z][a-z0-9-]*(?:\s+[a-z][a-z0-9-]*){0,3})"
        r"(?=\s+\d{1,3}(?:\.\d+)?\s*%?|$)",
        re.IGNORECASE
    )

    values = []
    for match in pattern.finditer(normalized):
        material = match.group("material").strip()
        if not material:
            continue

        # Prevent accidental capture of common section/identifier words.
        material = re.split(
            r"\b(?:shell|liner|lining|body|rn|ca|made|size|color|colour)\b.*$",
            material,
            maxsplit=1
        )[0].strip()

        if material:
            values.append(
                f"{match.group('pct')}% {material}"
            )

    return values


def normalize_composition(values):
    # Preserve source order. Ordering is presentation/semantic information;
    # sorting it hides which exact component differs.
    return [
        normalize_text(value)
        for value in values
        if normalize_text(value)
    ]


def analyze_content_difference(expected, actual):
    """Return the actual Content defect in QC-friendly language."""
    expected_values = extract_content_values(expected)
    actual_values = extract_content_values(actual)

    if not expected_values or not actual_values:
        return f"Expected: {expected} | Found: {actual}"

    issues = []
    from difflib import SequenceMatcher

    max_len = max(len(expected_values), len(actual_values))

    for index in range(max_len):
        exp = expected_values[index] if index < len(expected_values) else None
        act = actual_values[index] if index < len(actual_values) else None

        if exp is None:
            issues.append(f"Extra content in PDF: {act}")
            continue

        if act is None:
            issues.append(f"Missing from PDF: {exp}")
            continue

        exp_match = re.fullmatch(
            r"(?P<pct>\d+(?:\.\d+)?)%\s*(?P<material>.+)",
            normalize_text(exp),
            re.IGNORECASE
        )
        act_match = re.fullmatch(
            r"(?P<pct>\d+(?:\.\d+)?)%\s*(?P<material>.+)",
            normalize_text(act),
            re.IGNORECASE
        )

        if not exp_match or not act_match:
            if normalize_text(exp) != normalize_text(act):
                issues.append(f"Content mismatch: {exp} → {act}")
            continue

        exp_pct = exp_match.group("pct")
        act_pct = act_match.group("pct")
        exp_material = exp_match.group("material").strip()
        act_material = act_match.group("material").strip()

        if exp_pct != act_pct:
            issues.append(
                f"Percentage mismatch: Order Form {exp_pct}% → PDF {act_pct}% ({exp_material.upper()})"
            )
            continue

        if normalize_text(exp_material) == normalize_text(act_material):
            continue

        similarity = SequenceMatcher(
            None,
            normalize_text(exp_material),
            normalize_text(act_material),
            autojunk=False
        ).ratio()

        if similarity >= 0.70:
            issues.append(
                f'Spelling mistake: Order Form "{exp_material.upper()}" → PDF "{act_material.upper()}"'
            )
        else:
            issues.append(
                f'Material mismatch: Order Form "{exp_material.upper()}" → PDF "{act_material.upper()}"'
            )

    return "; ".join(issues) if issues else "Content differs."


def composition_matches(expected, actual):
    expected_values = normalize_composition(
        extract_content_values(expected)
    )
    actual_values = normalize_composition(
        extract_content_values(actual)
    )
    return bool(expected_values and expected_values == actual_values)


# =========================================================
# FIELD-SPECIFIC REGIONS
# =========================================================

FIELD_ANCHORS = {
    "COO": ["made in", "fabrique en", "hecho en"],
    "CARE": ["machine wash", "wash", "laver", "laver", "lavar", "bleach", "dry clean"],
    "CONTENT": ["%", "shell", "liner", "lining", "body", "fiber", "fibre", "content", "composition"],
    "RN": ["rn", "ca"],
    "IDENTIFIER": ["sku", "item code", "item no", "item number", "style", "style code"],
    "SIZE": ["size"],
    "COLOR": ["color", "colour"],
    "GENDER": ["boys", "girls", "women", "men", "unisex"],
    "BATCH": ["batch", "lot"],
    "QUANTITY": ["quantity", "qty", "units", "pcs"],
    "BRAND": ["brand"],
    "ATTRIBUTE": ["attribute", "technology", "feature"],
    "SYMBOL": [],
    "OSZ": [],
    "GENERAL": [],
}


def line_relevant_for_field(line_text, field_name):
    field_type = get_field_type(field_name)
    normalized = normalize_text(line_text)

    if not normalized:
        return False

    if field_type == "SYMBOL":
        return False

    if field_type == "OSZ":
        # OSZ values are resolved from a sequence, not generic relevance.
        return False

    for anchor in FIELD_ANCHORS.get(field_type, []):
        anchor_norm = normalize_text(anchor)
        if anchor_norm and anchor_norm in normalized:
            return True

    region = get_field_region(field_name)

    region_markers = {
        "EN": ["made in", "machine wash", "shell", "liner", "polyester", "bleach"],
        "FR": ["fabrique en", "laver", "extérieur", "doublure", "polyester", "sans chlore"],
        "SP": ["hecho en", "lavar", "forro", "poliester", "cloro"],
    }

    for marker in region_markers.get(region, []):
        if normalize_text(marker) in normalized:
            return True

    return False


def collect_region_from_anchor(state, field_name):
    """Collect a meaningful field region without crossing another major field."""
    lines = state["lines"]
    field_type = get_field_type(field_name)

    start_idx = None

    for idx, line in enumerate(lines):
        if not line_is_available(line, state):
            continue
        if line_relevant_for_field(line["text"], field_name):
            start_idx = idx
            break

    if start_idx is None:
        return None

    region = []

    stop_types = {
        "COO": ("RN", "CONTENT", "CARE"),
        "CONTENT": ("COO", "RN", "CARE"),
        "CARE": ("COO", "CONTENT", "RN"),
    }

    stop_patterns = {
        "COO": [r"\b(?:rn|ca)\s*[#:.-]?\s*\w+"],
        "CONTENT": [r"\b(?:made in|rn|ca)\b"],
        "CARE": [r"\b(?:made in|rn|ca)\b"],
    }

    for idx in range(start_idx, len(lines)):
        line = lines[idx]

        if not line_is_available(line, state):
            break

        text = line["text"]
        norm = line["norm"]

        if idx > start_idx:
            if field_type in stop_types:
                if any(
                    any(marker in norm for marker in FIELD_ANCHORS.get(stop_type, []))
                    for stop_type in stop_types[field_type]
                ):
                    break

            if field_type in stop_patterns:
                if any(re.search(pattern, norm) for pattern in stop_patterns[field_type]):
                    break

        region.append(line)

        # Avoid swallowing an entire page.
        if len(region) >= 20:
            break

    return region or None


# =========================================================
# OSZ SEQUENCE DETECTION
# =========================================================

def get_osz_index(field_name):
    compact = normalize_text(field_name).replace(" ", "")
    match = re.fullmatch(r"osz(\d+)", compact)
    return int(match.group(1)) if match else None


def extract_standalone_numeric_runs(state):
    runs = []
    current = []

    for line in state["lines"]:
        if not line_is_available(line, state):
            if current:
                runs.append(current)
                current = []
            continue

        norm = normalize_text(line["text"])

        # Standalone integer only. This deliberately excludes RN# 55285,
        # dates, dimensions, item codes and mixed text.
        if re.fullmatch(r"[-+]?\d+", norm):
            current.append(line)
        else:
            if current:
                runs.append(current)
                current = []

    if current:
        runs.append(current)

    return [run for run in runs if run]


def find_osz_value(
    expected,
    field_name,
    state,
    osz_group_size=1
):
    expected_num = normalize_numeric(expected)
    index = get_osz_index(field_name)

    if expected_num is None or index is None:
        return None

    # Explicit OSZ label, when available.
    label_pattern = re.compile(
        rf"\bosz\s*{index}\s*[:#=-]?\s*(\d+)\b",
        re.IGNORECASE
    )

    for line in state["lines"]:
        if not line_is_available(line, state):
            continue
        match = label_pattern.search(line["text"])
        if not match:
            continue

        actual = normalize_numeric(match.group(1))
        consume_lines(state, [line])
        if actual == expected_num:
            return {
                "status": "PASS",
                "pdf": line["text"],
                "difference": "—",
                "match_type": "OSZ_LABEL"
            }
        return {
            "status": "FAIL",
            "pdf": line["text"],
            "difference": f"Expected: {expected} | Found: {match.group(1)}",
            "match_type": "OSZ_LABEL_MISMATCH"
        }

    # Coordinate-aware OSZ mapping is preferred because OCR can merge a
    # standalone number with an unrelated neighbouring line (for example
    # turning a clean sequence 6 / 8 into a line such as "0 8").
    candidates = [
        candidate
        for candidate in state.get("osz_sequence_candidates", [])
        if len(candidate.get("items", [])) >= index
    ]

    if candidates:
        # First choose a candidate whose index already agrees with the Excel
        # value. This prevents an unrelated numeric run elsewhere on the page
        # from being selected merely because it is long.
        matching_candidates = [
            candidate
            for candidate in candidates
            if _osz_candidate_matches_expected(candidate, index, expected_num)
        ]
        chosen = max(
            matching_candidates or candidates,
            key=lambda candidate: candidate.get("score", 0.0)
        )

        item = chosen["items"][index - 1]
        actual = normalize_numeric(item.get("value"))
        actual_text = str(item.get("value", "")).strip() or "Not found"

        if actual == expected_num:
            return {
                "status": "PASS",
                "pdf": actual_text,
                "difference": "—",
                "match_type": "OSZ_SEQUENCE_GEOMETRY"
            }

        if actual is not None:
            return {
                "status": "FAIL",
                "pdf": actual_text,
                "difference": f"Expected: {expected} | Found: {actual_text}",
                "match_type": "OSZ_SEQUENCE_GEOMETRY_MISMATCH"
            }

    # Original line-based sequence logic remains as the final fallback.
    # This preserves the behaviour for PDFs whose text layer already exposes
    # one clean standalone number per line.
    runs = [run for run in state.get("osz_runs", []) if len(run) >= index]
    if not runs:
        return None

    runs.sort(key=lambda run: (-len(run), run[0]["line_id"]))
    run = runs[0]
    line = run[index - 1]

    if not line_is_available(line, state):
        return None

    actual = normalize_numeric(line["text"])

    if actual == expected_num:
        consume_lines(state, [line])
        return {
            "status": "PASS",
            "pdf": line["text"],
            "difference": "—",
            "match_type": "OSZ_SEQUENCE"
        }

    if actual is not None:
        consume_lines(state, [line])
        return {
            "status": "FAIL",
            "pdf": line["text"],
            "difference": f"Expected: {expected} | Found: {line['text']}",
            "match_type": "OSZ_SEQUENCE_MISMATCH"
        }

    return None


# =========================================================
# DIFFERENCE DESCRIPTION
# =========================================================

def describe_text_difference(expected, actual):
    """Explanation only; never used as a PASS/FAIL decision."""
    expected_tokens = tokenize(expected)
    actual_tokens = tokenize(actual)

    if not expected_tokens:
        return "Expected value is blank."
    if not actual_tokens:
        return "Expected value is missing from the output."

    from difflib import SequenceMatcher

    matcher = SequenceMatcher(
        None,
        expected_tokens,
        actual_tokens,
        autojunk=False
    )

    differences = []

    for tag, a1, a2, b1, b2 in matcher.get_opcodes():
        if tag == "equal":
            continue

        left = " ".join(expected_tokens[a1:a2]).strip()
        right = " ".join(actual_tokens[b1:b2]).strip()

        if tag == "replace":
            differences.append(f"{left} → {right}")
        elif tag == "delete":
            differences.append(f"Missing: {left}")
        elif tag == "insert":
            differences.append(f"Extra: {right}")

    return "; ".join(differences[:8]) or "Content differs."


# =========================================================
# FIELD CHECKERS
# =========================================================

def find_care_region(state, field_name):
    region = get_field_region(field_name)
    marker_sets = {
        "EN": ["machine wash", "wash", "bleach", "dry clean", "tumble dry", "cool iron"],
        "FR": ["laver", "blanchiment", "nettoyage", "sécher", "repasser"],
        "SP": ["lavar", "cloro", "secadora", "plancha", "limpieza en seco"],
        "": ["machine wash", "wash", "bleach", "dry clean", "laver", "lavar"],
    }
    markers = [normalize_text(m) for m in marker_sets.get(region, marker_sets[""]) if normalize_text(m)]

    lines = state["lines"]
    start = None
    for idx, line in enumerate(lines):
        if not line_is_available(line, state):
            continue
        if any(marker in line["norm"] for marker in markers):
            start = idx
            break

    if start is None:
        return None

    region_lines = []
    for idx in range(start, len(lines)):
        line = lines[idx]
        if not line_is_available(line, state):
            break

        if idx > start:
            if extract_coo_value(line["text"]) or extract_rn_value(line["text"]):
                break
            if extract_content_values(line["text"]):
                break

            # A non-text/symbol line, an isolated number, or a new obvious
            # technical line ends the care region. This is essential so symbols,
            # RN and OSZ data are not swallowed by the care matcher.
            ascii_letters = len(re.findall(r"[A-Za-z]", line["text"]))
            if ascii_letters == 0:
                break

            if re.fullmatch(r"\s*[-+]?\d+(?:\.\d+)?\s*", line["text"]):
                break

        region_lines.append(line)
        if len(region_lines) >= 20:
            break

    return region_lines or None


def check_field(
    expected,
    field_name,
    state,
    osz_group_size=1
):
    """Deterministic field validation. PASS is always attempted first."""

    if is_blank_value(expected):
        return {
            "status": "SKIP",
            "pdf": "—",
            "difference": "Blank Order Form value — field ignored.",
            "match_type": "BLANK"
        }

    expected = str(expected).strip()
    field_type = get_field_type(field_name)

    # -----------------------------------------------------
    # OSZ sequence
    # -----------------------------------------------------
    if field_type == "OSZ":
        result = find_osz_value(
            expected,
            field_name,
            state,
            osz_group_size=osz_group_size
        )
        if result:
            return result

        return {
            "status": "NOT FOUND",
            "pdf": "Not found",
            "difference": "OSZ sequence/value was not detected.",
            "match_type": "NOT_FOUND"
        }

    # -----------------------------------------------------
    # Symbol: exact symbol/text only. Never substring-match.
    # -----------------------------------------------------
    if field_type == "SYMBOL":
        for line in state["lines"]:
            if not line_is_available(line, state):
                continue

            if normalize_symbol_text(expected) == normalize_symbol_text(line["text"]):
                consume_lines(state, [line])
                return {
                    "status": "PASS",
                    "pdf": line["text"],
                    "difference": "—",
                    "match_type": "SYMBOL_EXACT"
                }

        return {
            "status": "NOT FOUND",
            "pdf": "Not found",
            "difference": "Symbol/value was not detected exactly.",
            "match_type": "NOT_FOUND"
        }

    # -----------------------------------------------------
    # CONTENT: parse the composition before any generic text check.
    # -----------------------------------------------------
    if field_type == "CONTENT":
        expected_values = normalize_composition(
            extract_content_values(expected)
        )

        if expected_values:
            # Search contiguous windows, but only over unconsumed lines.
            available = [
                line for line in state["lines"]
                if line_is_available(line, state)
            ]

            max_window = min(8, len(available))

            for size in range(1, max_window + 1):
                for start in range(0, len(available) - size + 1):
                    candidate = available[start:start + size]
                    ids = [line["line_id"] for line in candidate]
                    if ids != list(range(ids[0], ids[-1] + 1)):
                        continue

                    text = join_lines(candidate)
                    actual_values = normalize_composition(
                        extract_content_values(text)
                    )

                    if expected_values == actual_values:
                        consume_lines(state, candidate)
                        return {
                            "status": "PASS",
                            "pdf": text,
                            "difference": "—",
                            "match_type": "CONTENT_EXACT"
                        }

            # If we have a relevant composition region with overlapping
            # components, report the actual composition as FAIL.
            for line in state["lines"]:
                if not line_is_available(line, state):
                    continue
                text = line["text"]
                actual_values = normalize_composition(
                    extract_content_values(text)
                )
                if not actual_values:
                    continue
                if set(expected_values) & set(actual_values):
                    consume_lines(state, [line])
                    return {
                        "status": "FAIL",
                        "pdf": text,
                        "difference": (
                            f"Expected: {expected} | "
                            f"Found: {' | '.join(actual_values)}"
                        ),
                        "match_type": "CONTENT_MISMATCH"
                    }

            return {
                "status": "NOT FOUND",
                "pdf": "Not found",
                "difference": "Expected composition was not detected.",
                "match_type": "NOT_FOUND"
            }

    # -----------------------------------------------------
    # RN
    # -----------------------------------------------------
    if field_type == "RN":
        expected_rn = extract_rn_value(expected)
        expected_target = normalize_text(
            expected_rn if expected_rn else expected
        )

        for line in state["lines"]:
            if not line_is_available(line, state):
                continue

            actual_rn = extract_rn_value(line["text"])
            if actual_rn is None:
                continue

            if normalize_text(actual_rn) == expected_target:
                consume_lines(state, [line])
                return {
                    "status": "PASS",
                    "pdf": line["text"],
                    "difference": "—",
                    "match_type": "RN_EXACT"
                }

            consume_lines(state, [line])
            return {
                "status": "FAIL",
                "pdf": line["text"],
                "difference": (
                    f"Expected: {expected} | "
                    f"Found: RN# {actual_rn}"
                ),
                "match_type": "RN_MISMATCH"
            }

        return {
            "status": "NOT FOUND",
            "pdf": "Not found",
            "difference": "RN value was not detected.",
            "match_type": "NOT_FOUND"
        }

    # -----------------------------------------------------
    # COO
    # -----------------------------------------------------
    if field_type == "COO":
        expected_coo = extract_coo_value(expected)
        expected_target = normalize_text(
            expected_coo if expected_coo else expected
        )
        expected_region = get_field_region(field_name)

        # Prefer the requested language. Do not let English COO satisfy French
        # COO merely because the country happens to be the same.
        candidates = []
        for line in state["lines"]:
            if not line_is_available(line, state):
                continue
            actual_coo = extract_coo_value(line["text"])
            if not actual_coo:
                continue
            region = coo_language(line["text"])
            candidates.append((line, actual_coo, region))

        preferred = [
            item for item in candidates
            if expected_region and item[2] == expected_region
        ]

        # When no language suffix is supplied, any COO language may be used.
        search_candidates = preferred if preferred else (
            candidates if not expected_region else []
        )

        for line, actual_coo, _region in search_candidates:
            if normalize_text(actual_coo) == expected_target:
                consume_lines(state, [line])
                return {
                    "status": "PASS",
                    "pdf": line["text"],
                    "difference": "—",
                    "match_type": "COO_EXACT"
                }

        # A same-language different COO is a genuine FAIL.
        for line, actual_coo, _region in search_candidates:
            consume_lines(state, [line])
            return {
                "status": "FAIL",
                "pdf": line["text"],
                "difference": (
                    f"Expected: {expected} | "
                    f"Found: {actual_coo}"
                ),
                "match_type": "COO_MISMATCH"
            }

        return {
            "status": "NOT FOUND",
            "pdf": "Not found",
            "difference": (
                "Expected COO was not detected in the requested language/region."
            ),
            "match_type": "NOT_FOUND"
        }

    # -----------------------------------------------------
    # GENDER
    # -----------------------------------------------------
    if field_type == "GENDER":
        expected_gender = extract_gender_value(expected)
        expected_target = normalize_text(
            expected_gender if expected_gender else expected
        )

        for line in state["lines"]:
            if not line_is_available(line, state):
                continue
            actual_gender = extract_gender_value(line["text"])
            if not actual_gender:
                continue

            if normalize_text(actual_gender) == expected_target:
                consume_lines(state, [line])
                return {
                    "status": "PASS",
                    "pdf": line["text"],
                    "difference": "—",
                    "match_type": "GENDER_EXACT"
                }

            consume_lines(state, [line])
            return {
                "status": "FAIL",
                "pdf": line["text"],
                "difference": (
                    f"Expected: {expected} | "
                    f"Found: {actual_gender}"
                ),
                "match_type": "GENDER_MISMATCH"
            }

        return {
            "status": "NOT FOUND",
            "pdf": "Not found",
            "difference": "Gender value was not detected.",
            "match_type": "NOT_FOUND"
        }

    # -----------------------------------------------------
    # SIZE
    # -----------------------------------------------------
    if field_type == "SIZE":
        expected_size = normalize_text(expected)
        expected_size = re.sub(
            r"^size\s*[:#-]?\s*",
            "",
            expected_size
        ).strip()

        # Prefer labeled size blocks.
        for line in state["lines"]:
            if not line_is_available(line, state):
                continue
            actual_size = extract_size_value(line["text"])
            if actual_size is None:
                continue

            if normalize_text(actual_size) == expected_size:
                consume_lines(state, [line])
                return {
                    "status": "PASS",
                    "pdf": line["text"],
                    "difference": "—",
                    "match_type": "SIZE_EXACT"
                }

            consume_lines(state, [line])
            return {
                "status": "FAIL",
                "pdf": line["text"],
                "difference": (
                    f"Expected: {expected} | "
                    f"Found: {actual_size}"
                ),
                "match_type": "SIZE_MISMATCH"
            }

        # Then exact text, useful when artwork prints only the value.
        exact = find_exact_lines(
            expected,
            field_name,
            state,
            max_window=1
        )
        if exact:
            lines, match_info = exact
            consume_match(state, match_info)
            pdf_value = match_info.get("actual") or join_lines(lines)
            return {
                "status": "PASS",
                "pdf": pdf_value,
                "difference": "—",
                "match_type": match_info.get("match_type", "EXACT")
            }

        return {
            "status": "NOT FOUND",
            "pdf": "Not found",
            "difference": "Size value was not detected.",
            "match_type": "NOT_FOUND"
        }

    # -----------------------------------------------------
    # COLOR
    # -----------------------------------------------------
    if field_type == "COLOR":
        expected_color = normalize_text(expected)
        expected_color = re.sub(
            r"^(?:color|colour)\s*[:#-]?\s*",
            "",
            expected_color
        ).strip()

        for line in state["lines"]:
            if not line_is_available(line, state):
                continue
            actual_color = extract_color_value(line["text"])
            if actual_color is None:
                continue

            if normalize_text(actual_color) == expected_color:
                consume_lines(state, [line])
                return {
                    "status": "PASS",
                    "pdf": line["text"],
                    "difference": "—",
                    "match_type": "COLOR_EXACT"
                }

            consume_lines(state, [line])
            return {
                "status": "FAIL",
                "pdf": line["text"],
                "difference": (
                    f"Expected: {expected} | "
                    f"Found: {actual_color}"
                ),
                "match_type": "COLOR_MISMATCH"
            }

        return _generic_exact_field(expected, field_name, state)

    # -----------------------------------------------------
    # CARE
    # -----------------------------------------------------
    if field_type == "CARE":
        care_region = find_care_region(state, field_name)

        if care_region:
            actual_text = join_lines(care_region)
            expected_norm = normalize_text(expected)
            actual_norm = normalize_text(actual_text)

            if expected_norm == actual_norm or expected_norm in actual_norm:
                consume_lines(state, care_region)
                return {
                    "status": "PASS",
                    "pdf": actual_text,
                    "difference": "—",
                    "match_type": "CARE_EXACT_REGION"
                }

            expected_tokens = set(tokenize(expected))
            actual_tokens = set(tokenize(actual_text))
            common = expected_tokens & actual_tokens

            if expected_tokens and len(common) >= max(3, int(len(expected_tokens) * 0.55)):
                consume_lines(state, care_region)
                return {
                    "status": "FAIL",
                    "pdf": actual_text,
                    "difference": describe_text_difference(expected, actual_text),
                    "match_type": "CARE_MISMATCH"
                }

        return {
            "status": "NOT FOUND",
            "pdf": "Not found",
            "difference": "Care instruction was not detected in the relevant artwork region.",
            "match_type": "NOT_FOUND"
        }

    # -----------------------------------------------------
    # IDENTIFIER
    # -----------------------------------------------------
    if field_type == "IDENTIFIER":
        identifier = find_identifier_match(expected, state)
        if identifier:
            consume_span(
                state,
                identifier["line"],
                identifier["start"],
                identifier["end"]
            )
            return {
                "status": identifier["kind"],
                "pdf": identifier["actual"],
                "difference": identifier["difference"],
                "match_type": identifier["match_type"]
            }

        return {
            "status": "NOT FOUND",
            "pdf": "Not found",
            "difference": "Expected identifier was not detected.",
            "match_type": "NOT_FOUND"
        }

    # -----------------------------------------------------
    # GENERAL / BRAND / ATTRIBUTE / BATCH / QUANTITY
    # -----------------------------------------------------
    return _generic_exact_field(
        expected,
        field_name,
        state
    )


def _generic_exact_field(expected, field_name, state):
    exact = find_exact_lines(
        expected,
        field_name,
        state,
        max_window=8
    )

    if exact:
        lines, match_info = exact
        consume_match(state, match_info)

        if match_info.get("actual"):
            pdf_value = match_info["actual"]
        else:
            pdf_value = join_lines(lines)

        return {
            "status": "PASS",
            "pdf": pdf_value,
            "difference": "—",
            "match_type": match_info.get("match_type", "EXACT")
        }

    # IMPORTANT: unknown/general fields do not generate invented FAILs.
    return {
        "status": "NOT FOUND",
        "pdf": "Not found",
        "difference": "Expected value was not detected.",
        "match_type": "NOT_FOUND"
    }


# =========================================================
# FIELD MATCHING ORDER
# =========================================================

FIELD_PRIORITY = {
    "OSZ": 10,
    "RN": 20,
    "IDENTIFIER": 20,
    "COO": 30,
    "CONTENT": 40,
    "CARE": 50,
    "GENDER": 60,
    "COLOR": 60,
    "SIZE": 60,
    "BATCH": 60,
    "QUANTITY": 60,
    "SYMBOL": 70,
    "BRAND": 80,
    "ATTRIBUTE": 90,
    "GENERAL": 100,
}


def order_fields_for_matching(fields):
    indexed = list(enumerate(fields))

    return [
        field
        for _index, field in sorted(
            indexed,
            key=lambda pair: (
                FIELD_PRIORITY.get(
                    get_field_type(pair[1]),
                    100
                ),
                pair[0]
            )
        )
    ]


# =========================================================
# BUILD REPORT
# =========================================================

def build_report(
    df,
    pdf_pages,
    selected_fields,
    product_type,
    page_row_mapping=None
):
    """
    Full validation report.

    Matching order is optimized for specificity, but report output is returned
    in the original selected-field order per page.
    """

    results = []
    field_no = 1

    # Pre-compute OSZ group sizes per row. They are used only to make sequence
    # mapping safe when more than one OSZ field exists.
    osz_fields = [
        field for field in selected_fields
        if get_field_type(field) == "OSZ"
    ]
    osz_group_size = len(osz_fields)

    for page_index, page in enumerate(pdf_pages):

        page_number = int(page.get("page", page_index + 1))

        if page_row_mapping is not None and page_number not in page_row_mapping:
            continue

        if page_row_mapping and page_number in page_row_mapping:
            excel_index = int(page_row_mapping[page_number])
        else:
            excel_index = page_index

        if excel_index >= len(df):
            for field in selected_fields:
                results.append({
                    "FIELD NO": field_no,
                    "PDF PAGE": page_number,
                    "EXCEL ROW": "N/A",
                    "FIELD": field,
                    "ORDER FORM DATA": "No Excel row",
                    "PDF OUTPUT": "No corresponding Order Form row",
                    "STATUS": "NOT FOUND",
                    "DIFFERENCE": "No corresponding Excel row."
                })
                field_no += 1
            continue

        row = df.iloc[excel_index]
        state = build_page_state(
            page,
            product_type
        )

        # Match in smart specificity order so a precise RN/OSZ/COO matcher gets
        # the relevant artwork before a broad textual field can consume it.
        matching_fields = order_fields_for_matching(
            selected_fields
        )

        page_results = {}

        for field in matching_fields:

            value = "" if is_blank_value(row[field]) else str(row[field]).strip()

            result = check_field(
                value,
                field,
                state,
                osz_group_size=osz_group_size
            )

            page_results[field] = {
                "field": field,
                "value": value,
                "result": result
            }

        # Return results in the user's original field order, not matching order.
        for field in selected_fields:
            item = page_results[field]
            result = item["result"]

            results.append({
                "FIELD NO": field_no,
                "PDF PAGE": page_number,
                "EXCEL ROW": excel_index + 2,
                "FIELD": field,
                "ORDER FORM DATA": item["value"],
                "PDF OUTPUT": result["pdf"],
                "STATUS": result["status"],
                "DIFFERENCE": result["difference"]
            })

            field_no += 1

    return pd.DataFrame(results)


# =========================================================
# AUTO DETECT
# =========================================================

def _auto_detect_page_text(page):
    """Return normalized OCR/direct text for one artwork page."""
    if not page:
        return ""

    values = []
    # Prefer OCR because the artwork may be a scanned/non-editable PDF.
    for key in ("ocr_text", "ocr_alt_text", "text", "direct_text"):
        value = page.get(key, "")
        if value and str(value).strip():
            values.append(str(value))

    return "\n".join(values)


def _auto_detect_lines(page):
    text = _auto_detect_page_text(page)
    return [line for line in (normalize_text(x) for x in str(text).splitlines()) if line]


def _auto_number_evidence(expected, text):
    """Exact numeric evidence with boundaries; never accepts a substring."""
    numeric = normalize_numeric(expected)
    if numeric is None:
        return False
    return bool(
        re.search(
            rf"(?<![A-Za-z0-9]){re.escape(numeric)}(?![A-Za-z0-9])",
            normalize_text(text)
        )
    )


def _auto_identifier_evidence(expected, field_name, text):
    """Asymmetric identifier evidence used for item/style/supplier codes."""
    expected_compact = re.sub(r"[^a-z0-9]", "", normalize_text(expected))
    field_compact = (
        normalize_text(field_name)
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )

    # Supplier/vendor IDs such as USX are intentionally allowed at 3 chars
    # because the artwork may contain an appended static suffix (USX609).
    minimum_length = 3 if any(
        key in field_compact
        for key in ("supwsp", "supplier", "vendorid", "vendorcode")
    ) else 4

    if not expected_compact or len(expected_compact) < minimum_length:
        return False

    tokens = re.findall(r"[a-z0-9]+", normalize_text(text))
    for token in tokens:
        if token == expected_compact:
            return True
        if len(expected_compact) >= 5 and token.startswith(expected_compact):
            return True
        if len(expected_compact) == 3 and token.startswith(expected_compact):
            return True

    return False


def _auto_material_evidence(expected, text):
    """Strong evidence for canonical visible composition-description fields."""
    expected_parts = extract_content_values(expected)
    if not expected_parts:
        return False

    norm_text = normalize_text(text)
    compact_text_value = re.sub(r"[^a-z0-9%]", "", norm_text)

    matches = 0
    for part in expected_parts:
        part_norm = normalize_text(part)
        material = re.sub(r"[^a-z]", "", part_norm.casefold())
        pct_match = re.search(r"(\d+(?:\.\d+)?)\s*%?", part_norm)

        if not material or len(material) < 3:
            continue

        if material in compact_text_value:
            if pct_match and re.search(
                rf"{re.escape(pct_match.group(1))}\s*%?\s*{re.escape(material)}",
                norm_text,
            ):
                matches += 1
            elif re.search(rf"\b{re.escape(material)}\b", norm_text):
                matches += 1

    required = 2 if len(expected_parts) >= 2 else 1
    has_composition_shape = bool(
        re.search(r"\d+(?:\.\d+)?\s*%?", norm_text)
        and re.search(r"[a-z]{3,}", norm_text)
    )
    return matches >= required and has_composition_shape


def _auto_coo_evidence(expected, field_name, text):
    """High-confidence COO evidence; codes like MADE_IN=F are not artwork text."""
    from rapidfuzz import fuzz

    expected_coo = extract_coo_value(expected)
    target = normalize_text(expected_coo if expected_coo else expected)

    # One/two-letter language/origin codes are internal codes, not visible COO.
    if re.fullmatch(r"[a-z]{1,2}", target):
        return False

    target_compact = re.sub(r"[^a-z0-9]", "", target)
    if len(target_compact) < 4:
        return False

    for line in _auto_detect_lines({"ocr_text": text}):
        line_compact = re.sub(r"[^a-z0-9]", "", line)
        if target_compact == line_compact:
            return True
        if target_compact and target_compact in line_compact:
            return True
        if fuzz.ratio(target_compact, line_compact) >= 92:
            return True

    return False


def _auto_generic_field_allowed(field_name):
    """
    Allow only known artwork-variable semantics among GENERAL technical fields.
    This prevents operational/database columns from being auto-selected just
    because they are populated.
    """
    compact = (
        normalize_text(field_name)
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )

    patterns = (
        "stylewofinish",
        "cdstyle",
        "cdfinishing",
        "finishing",
        "cdimport",
        "import",
        "designstyle",
        "lblstyle",
        "antfamily",
        "family",
        "compodsc",
        "lhcompodsc",
    )
    return any(token in compact for token in patterns)


def _auto_content_field_allowed(field_name):
    """Only canonical composition-description fields are auto-detected."""
    compact = (
        normalize_text(field_name)
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )
    return compact.startswith("compodsc") or compact.startswith("lhcompodsc")


def _auto_text_evidence(expected, field_name, text):
    """Conservative Auto Detect evidence check."""
    from rapidfuzz import fuzz

    expected_norm = normalize_text(expected)
    if not expected_norm:
        return False

    field_type = get_field_type(field_name)
    compact_expected = re.sub(r"[^a-z0-9]", "", expected_norm)
    norm_text = normalize_text(text)

    if field_type == "IDENTIFIER":
        return _auto_identifier_evidence(expected, field_name, norm_text)

    if field_type == "RN":
        digits = re.sub(r"\D", "", expected_norm)
        if not digits or len(digits) < 4:
            return False
        if re.search(r"\b(?:rn|ca)\s*[#:\-./ ]*\d+", norm_text):
            return digits in re.sub(r"\D", "", norm_text)
        return bool(re.search(rf"(?<!\d){re.escape(digits)}(?!\d)", norm_text))

    if field_type == "COO":
        return _auto_coo_evidence(expected, field_name, norm_text)

    if field_type == "CONTENT":
        return _auto_content_field_allowed(field_name) and _auto_material_evidence(expected, norm_text)

    if field_type == "CARE":
        tokens = [t for t in re.findall(r"[a-z]+", expected_norm) if len(t) >= 4]
        if len(tokens) < 3:
            return False
        unique = set(tokens)
        hits = sum(1 for token in unique if token in norm_text)
        return hits >= max(3, int(len(unique) * 0.35))

    if field_type == "SIZE":
        compact_text = re.sub(r"[^a-z0-9./-]", "", norm_text)
        if compact_expected and compact_expected in compact_text:
            return True
        return _auto_number_evidence(expected_norm, norm_text) and any(
            key in normalize_text(field_name).replace(" ", "")
            for key in ("size", "waist", "inseam", "alpha", "fit")
        )

    if field_type == "COLOR":
        if expected_norm in norm_text:
            return True
        lines = _auto_detect_lines({"ocr_text": text})
        if not lines or len(expected_norm) < 4:
            return False
        return max(fuzz.ratio(expected_norm, line) for line in lines) >= 86

    if field_type == "GENDER":
        aliases = {
            "men": ("men", "men's", "mens", "male"),
            "women": ("women", "women's", "womens", "female"),
            "boys": ("boys", "boy", "boy's", "boys'"),
            "girls": ("girls", "girl", "girl's", "girls'"),
        }
        key = expected_norm.replace("'", "")
        for base, variants in aliases.items():
            if key == base or key.rstrip("s") == base.rstrip("s"):
                return any(normalize_text(v) in norm_text for v in variants)
        return expected_norm in norm_text

    if field_type == "BRAND":
        if compact_expected and compact_expected in re.sub(r"[^a-z0-9]", "", norm_text):
            return True
        lines = _auto_detect_lines({"ocr_text": text})
        return bool(lines) and len(expected_norm) >= 4 and max(
            fuzz.ratio(expected_norm, line) for line in lines
        ) >= 88

    if field_type == "ATTRIBUTE":
        tokens = [t for t in re.findall(r"[a-z0-9]+", expected_norm) if len(t) >= 3]
        if not tokens:
            return False
        best = 0
        for line in _auto_detect_lines({"ocr_text": text}):
            line_tokens = set(re.findall(r"[a-z0-9]+", line))
            overlap = len(set(tokens) & line_tokens) / max(1, len(set(tokens)))
            best = max(best, overlap)
        return best >= 0.60

    if field_type in {"QUANTITY", "BATCH"}:
        # Quantities/lots are deliberately manual because artwork contains
        # many unrelated numbers and barcode data.
        return False

    if field_type == "OSZ":
        return _auto_number_evidence(expected, norm_text)

    # GENERAL: only selected semantic technical fields are eligible.
    if not _auto_generic_field_allowed(field_name):
        return False

    # Never auto-detect a one-letter technical code such as MADE_IN=F.
    if re.fullmatch(r"[a-z]{1,2}", expected_norm):
        return False

    if normalize_numeric(expected) is not None:
        return _auto_number_evidence(expected_norm, norm_text)

    lines = _auto_detect_lines({"ocr_text": text})
    if not lines:
        return False

    if compact_expected and compact_expected in re.sub(r"[^a-z0-9]", "", norm_text):
        return True

    if len(expected_norm) >= 6:
        return max(fuzz.ratio(expected_norm, line) for line in lines) >= 90

    return False


def _auto_candidate_priority(field_name):
    """Lower number = preferred canonical source column when duplicate values exist."""
    compact = (
        normalize_text(field_name)
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )

    if compact.startswith("compodsc"):
        return 0
    if compact.startswith("lhcompodsc"):
        return 10
    return 20


def _auto_field_signature(field_name, value):
    """Semantic duplicate signature so equivalent columns do not flood Auto Detect."""
    field_type = get_field_type(field_name)

    if field_type == "CONTENT":
        parts = extract_content_values(value)
        if parts:
            parsed = []
            for part in parts:
                match = re.match(
                    r"(\d+(?:\.\d+)?)%?\s*(.+)",
                    normalize_text(part)
                )
                if match:
                    parsed.append(
                        (
                            match.group(1),
                            normalize_text(match.group(2))
                        )
                    )
                else:
                    parsed.append(("", normalize_text(part)))
            return (field_type, tuple(parsed))

    return (field_type, normalize_text(value))


def auto_detect_fields(
    df,
    output_pages,
    product_type,
    page_row_mapping=None
):
    """
    Controlled Auto Detect.

    A field is selected only when:
      1. it is populated in the mapped Order Form row,
      2. its column belongs to an allowed artwork-variable family, and
      3. its actual value has strong evidence in the mapped artwork.

    Population alone is never enough.
    """
    available_fields = get_available_fields(df)
    allowed_types = {
        "IDENTIFIER",
        "BRAND",
        "GENDER",
        "SIZE",
        "COLOR",
        "COO",
        "CONTENT",
        "CARE",
        "ATTRIBUTE",
        "RN",
        "OSZ",
        "GENERAL",
    }

    candidates = []
    for field in available_fields:
        if is_admin_field(field):
            continue

        field_type = get_field_type(field)

        if field_type == "GENERAL" and not _auto_generic_field_allowed(field):
            continue

        if field_type == "CONTENT" and not _auto_content_field_allowed(field):
            continue

        if field_type not in allowed_types:
            continue

        compact = (
            normalize_text(field)
            .replace(" ", "")
            .replace("_", "")
            .replace("-", "")
        )

        # Never auto-select translated/internal material columns.
        if any(token in compact for token in (
            "p1mat",
            "multi",
            "translation",
            "greek",
            "arabic",
            "turkish",
            "indonesia",
            "matfull",
        )):
            if not compact.startswith("compodsc"):
                continue

        candidates.append(field)

    if not candidates or not output_pages:
        return []

    rows_to_check = []
    for page in output_pages:
        page_number = int(page.get("page", 1))

        if page_row_mapping is not None:
            if page_number not in page_row_mapping:
                continue
            row_index = int(page_row_mapping[page_number])
        elif len(df) == 1:
            row_index = 0
        else:
            row_index = page_number - 1

        if 0 <= row_index < len(df):
            rows_to_check.append((page, df.iloc[row_index]))

    if not rows_to_check:
        return []

    # Prefer canonical composition descriptions before equivalent helper columns.
    candidates = sorted(
        enumerate(candidates),
        key=lambda item: (_auto_candidate_priority(item[1]), item[0])
    )
    candidates = [field for _idx, field in candidates]

    detected = []
    detected_signatures = set()

    for field in candidates:
        found = False
        representative_value = None

        for page, row in rows_to_check:
            value = row.get(field, "")
            if is_blank_value(value):
                continue

            representative_value = str(value).strip()
            if _auto_text_evidence(
                value,
                field,
                _auto_detect_page_text(page)
            ):
                found = True
                break

        if not found or representative_value is None:
            continue

        signature = _auto_field_signature(field, representative_value)
        if signature in detected_signatures:
            continue

        detected_signatures.add(signature)
        detected.append(field)

    # Return fields in their original Excel order.
    original_order = {str(column): idx for idx, column in enumerate(df.columns)}
    detected.sort(key=lambda field: original_order.get(field, 10**9))
    return detected


# =========================================================
# STATUS COLORS
# =========================================================

def style_status(value):
    if value == "PASS":
        return (
            "background-color: #238636;"
            "color: white;"
            "font-weight: bold;"
        )

    if value == "FAIL":
        return (
            "background-color: #da3633;"
            "color: white;"
            "font-weight: bold;"
        )

    if value == "NOT FOUND":
        return (
            "background-color: #9e6a03;"
            "color: white;"
            "font-weight: bold;"
        )

    if value == "SKIP":
        return (
            "background-color: #555555;"
            "color: white;"
            "font-weight: bold;"
        )

    return ""



def create_excel_report(report, product_type, comparison_method, selected_fields, visual_pages=None):
    """Create a professional Excel QC report, including full artwork pages with highlights."""

    output = BytesIO()
    wb = Workbook()

    NAVY = "1F4E78"
    LIGHT_BLUE = "D9EAF7"
    GREEN = "238636"
    RED = "DA3633"
    AMBER = "9E6A03"
    GREY = "555555"
    WHITE = "FFFFFF"
    LIGHT_BORDER = "D9E1F2"

    pass_count = int((report["STATUS"] == "PASS").sum()) if not report.empty else 0
    fail_count = int((report["STATUS"] == "FAIL").sum()) if not report.empty else 0
    not_found_count = int((report["STATUS"] == "NOT FOUND").sum()) if not report.empty else 0
    skip_count = int((report["STATUS"] == "SKIP").sum()) if not report.empty else 0
    total_checks = len(report)

    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:F2")
    ws["A1"] = "PDF PROOFREADING QC REPORT"
    ws["A1"].font = Font(bold=True, size=20, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    summary = [
        ("Comparison Method", comparison_method or "—"),
        ("Product Type", product_type or "—"),
        ("Selected / Detected Fields", len(selected_fields or [])),
        ("Total Field Checks", total_checks),
        ("PASS", pass_count),
        ("FAIL", fail_count),
        ("NOT FOUND", not_found_count),
        ("IGNORED / SKIP", skip_count),
    ]

    for r, (label, value) in enumerate(summary, start=4):
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(r, 1).alignment = Alignment(vertical="center")
        ws.cell(r, 2).alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
    ws.freeze_panes = "A4"

    finding_row = 14
    ws.merge_cells(start_row=finding_row, start_column=1, end_row=finding_row, end_column=6)
    ws.cell(finding_row, 1, "QC NOTES")
    ws.cell(finding_row, 1).font = Font(bold=True, color=WHITE)
    ws.cell(finding_row, 1).fill = PatternFill("solid", fgColor=NAVY)

    notes = []
    if fail_count:
        notes.append(f"{fail_count} FAIL result(s) require review.")
    if not_found_count:
        notes.append(f"{not_found_count} field(s) could not be reliably located.")
    if skip_count:
        notes.append(f"{skip_count} blank/ignored field check(s) were skipped.")
    if not notes:
        notes.append("All checked fields passed.")
    for idx, note in enumerate(notes, start=finding_row + 1):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=6)
        ws.cell(idx, 1, "• " + note)
        ws.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical="top")

    comparison = wb.create_sheet("Field Comparison")
    display_report = add_visual_column(report, selected_fields)
    headers = list(display_report.columns)
    header_fill = PatternFill("solid", fgColor=NAVY)
    for col_idx, header in enumerate(headers, start=1):
        cell = comparison.cell(1, col_idx, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if "VISUAL" in headers:
        visual_header = comparison.cell(1, headers.index("VISUAL") + 1)
        visual_header.comment = None

    for row_idx, row in enumerate(display_report.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = comparison.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    status_col = None
    for idx, header in enumerate(headers, start=1):
        if header == "STATUS":
            status_col = idx
            break
    if status_col:
        for row_idx in range(2, comparison.max_row + 1):
            cell = comparison.cell(row_idx, status_col)
            status = str(cell.value or "")
            if status == "PASS":
                cell.fill = PatternFill("solid", fgColor=GREEN)
                cell.font = Font(color=WHITE, bold=True)
            elif status == "FAIL":
                cell.fill = PatternFill("solid", fgColor=RED)
                cell.font = Font(color=WHITE, bold=True)
            elif status == "NOT FOUND":
                cell.fill = PatternFill("solid", fgColor=AMBER)
                cell.font = Font(color=WHITE, bold=True)
            elif status == "SKIP":
                cell.fill = PatternFill("solid", fgColor=GREY)
                cell.font = Font(color=WHITE, bold=True)

    visual_col = headers.index("VISUAL") + 1 if "VISUAL" in headers else None
    if visual_col:
        field_col = headers.index("FIELD") + 1
        field_colors = get_field_visual_colors(selected_fields)
        for row_idx in range(2, comparison.max_row + 1):
            field_name = str(comparison.cell(row_idx, field_col).value or "")
            color = field_colors.get(field_name, "6B7280").lstrip("#").upper()
            cell = comparison.cell(row_idx, visual_col)
            cell.value = ""
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    width_map = {
        "FIELD NO": 12,
        "PDF PAGE": 12,
        "EXCEL ROW": 12,
        "FIELD": 24,
        "ORDER FORM DATA": 42,
        "PDF OUTPUT": 52,
        "STATUS": 16,
        "VISUAL": 12,
        "DIFFERENCE": 58,
        "MATCH TYPE": 22,
    }
    for col_idx, header in enumerate(headers, start=1):
        comparison.column_dimensions[get_column_letter(col_idx)].width = width_map.get(header, 20)
    comparison.row_dimensions[1].height = 32
    for row_idx in range(2, comparison.max_row + 1):
        comparison.row_dimensions[row_idx].height = 55
    comparison.freeze_panes = "A2"
    comparison.auto_filter.ref = comparison.dimensions

    thin = Side(style="thin", color=LIGHT_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in comparison.iter_rows():
        for cell in row:
            cell.border = border

    if visual_pages:
        visual = wb.create_sheet("Artwork Visual Validation")
        visual.column_dimensions["A"].width = 22
        visual.column_dimensions["B"].width = 24
        visual.column_dimensions["C"].width = 24
        visual["A1"] = "ARTWORK VISUAL VALIDATION"
        visual["A1"].font = Font(bold=True, size=16, color=WHITE)
        visual["A1"].fill = PatternFill("solid", fgColor=NAVY)
        visual.merge_cells("A1:C2")
        visual["A1"].alignment = Alignment(horizontal="center", vertical="center")

        row_cursor = 4
        for page in visual_pages:
            page_num = page.get("page")
            page_report = report[report["PDF PAGE"] == page_num] if "PDF PAGE" in report.columns else report.iloc[0:0]
            field_colors = get_field_visual_colors(selected_fields)
            highlighted = build_highlighted_page_image(
                page,
                page_report,
                field_colors=field_colors,
            )
            if highlighted is None:
                continue
            visual.cell(row_cursor, 1, f"Artwork Page {page_num}")
            visual.cell(row_cursor, 1).font = Font(bold=True, size=13, color=WHITE)
            visual.cell(row_cursor, 1).fill = PatternFill("solid", fgColor=NAVY)
            visual.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
            row_cursor += 1
            image_data = _visual_image_bytes(highlighted)
            if image_data:
                img = XLImage(BytesIO(image_data))
                img.width = min(560, highlighted.width)
                img.height = int(highlighted.height * (img.width / highlighted.width))
                anchor = f"A{row_cursor}"
                visual.add_image(img, anchor)
                row_height_count = max(20, int(img.height / 1.35))
                for r in range(row_cursor, row_cursor + max(1, int(row_height_count / 15))):
                    visual.row_dimensions[r].height = 15
                row_cursor += max(35, int(img.height / 14))
            row_cursor += 2

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def main():
    """Render Tool 1: Order Form → Output Check."""

    _apply_tool_css()

    # =========================================================
    # TOOL 1 SESSION STATE
    # =========================================================
    if "of_reset_id" not in st.session_state:
        st.session_state["of_reset_id"] = 0

    if "of_report" not in st.session_state:
        st.session_state["of_report"] = None

    if "of_visual_pages" not in st.session_state:
        st.session_state["of_visual_pages"] = None

    if "of_auto_detected_fields" not in st.session_state:
        st.session_state["of_auto_detected_fields"] = []

    if "of_auto_detect_key" not in st.session_state:
        st.session_state["of_auto_detect_key"] = None

    if "of_auto_output_pages" not in st.session_state:
        st.session_state["of_auto_output_pages"] = None

    # A code update must never leave an old comparison report visible in the
    # same Streamlit session. This is presentation/session hygiene only; it
    # does not modify any comparison decision.
    if st.session_state.get("of_report_build_version") != AUTO_DETECT_ENGINE_VERSION:
        st.session_state["of_report"] = None
        st.session_state["of_visual_pages"] = None
        st.session_state["of_report_selected_fields"] = []
        st.session_state["of_report_product_type"] = None
        st.session_state["of_report_comparison_method"] = None
        st.session_state["of_report_build_version"] = AUTO_DETECT_ENGINE_VERSION

    # =========================================================
    # TITLE
    # =========================================================
    st.markdown(
        '<div class="main-title">🔍 PDF Proofreader</div>',
        unsafe_allow_html=True
    )

    st.markdown(
        '<div class="sub-title">'
        'Compare selected variable Order Form fields against PDF artwork.'
        '</div>',
        unsafe_allow_html=True
    )

    # =========================================================
    # TOOL NAVIGATION
    # =========================================================
    nav_left, nav_right = st.columns([1, 1])

    with nav_left:
        if st.button("← HOME", key="of_back_home", width="stretch"):
            st.session_state["selected_tool"] = None
            st.session_state["of_report"] = None
            st.session_state["of_visual_pages"] = None
            st.session_state["of_auto_detected_fields"] = []
            st.session_state["of_auto_detect_key"] = None
            st.session_state["of_auto_output_pages"] = None
            st.session_state["of_report_build_version"] = None
            st.rerun()

    with nav_right:
        if st.button(
            "🆕 NEW START",
            key="of_new_start",
            width="stretch"
        ):
            st.session_state["of_reset_id"] += 1
            st.session_state["of_report"] = None
            st.session_state["of_visual_pages"] = None
            st.session_state["of_auto_detected_fields"] = []
            st.session_state["of_auto_detect_key"] = None
            st.session_state["of_auto_output_pages"] = None
            st.session_state["of_report_build_version"] = None
            st.rerun()

    # =========================================================
    # PRODUCT TYPE
    # =========================================================
    product_type = st.selectbox(
        "Select Product Type",
        options=[
            "----- SELECT -----",
            "PFL",
            "HTL",
            "Other"
        ],
        index=0,
        key=f"of_product_type_{st.session_state['of_reset_id']}",
        help=(
            "PFL = panelled artwork where variable data may continue "
            "across panels. HTL / Other = standard continuous-data comparison."
        )
    )

    if product_type == "----- SELECT -----":
        st.info(
            "Please select a Product Type before starting the comparison."
        )

    # =========================================================
    # UPLOAD AREA
    # =========================================================
    left_column, right_column = st.columns(2)

    with left_column:
        st.markdown(
            '<div class="section-title">📊 Order Form</div>',
            unsafe_allow_html=True
        )

        excel_file = st.file_uploader(
            "Upload Excel Order Form",
            type=["xlsx", "xls"],
            key=f"excel_upload_{st.session_state['of_reset_id']}"
        )

    with right_column:
        st.markdown(
            '<div class="section-title">📄 Output Artwork</div>',
            unsafe_allow_html=True
        )

        output_file = st.file_uploader(
            "Upload Output Artwork",
            type=["pdf", "jpg", "jpeg", "png"],
            key=f"output_upload_{st.session_state['of_reset_id']}"
        )

    # =========================================================
    # LOAD ORDER FORM
    # =========================================================
    df = None

    if excel_file:
        try:
            df = load_excel(excel_file)
        except Exception as error:
            st.error(
                f"Unable to read the Excel Order Form: {error}"
            )
            return

    # =========================================================
    # FILE INFORMATION + PAGE SELECTION + ROW MAPPING
    # =========================================================
    output_page_count = 0
    selected_pdf_pages = []
    page_row_mapping = {}
    mapping_ready = False

    if excel_file and output_file:
        try:
            output_page_count = get_output_page_count(output_file)
        except Exception as error:
            st.error(f"Unable to determine Output Artwork page count: {error}")
            output_page_count = 0

        st.markdown(
            '<div class="section-title">📌 File Information</div>',
            unsafe_allow_html=True
        )

        info1, info2, info3 = st.columns(3)
        with info1:
            st.metric("Excel Data Rows", len(df))
        with info2:
            st.metric("Output Pages", output_page_count)
        with info3:
            extension = str(output_file.name).split(".")[-1].upper()
            st.metric("Output Type", extension)

        if output_page_count == 1:
            selected_pdf_pages = [1]
            st.caption("One Output page detected. Select which Order Form data row this page should use.")
        elif output_page_count > 1:
            page_mode = st.radio(
                "Artwork page selection",
                options=["All Pages", "Specific Page(s)"],
                horizontal=True,
                key=f"page_mode_{st.session_state['of_reset_id']}"
            )

            if page_mode == "All Pages":
                selected_pdf_pages = list(range(1, output_page_count + 1))
            else:
                selected_pdf_pages = st.multiselect(
                    "Select Output Page(s)",
                    options=list(range(1, output_page_count + 1)),
                    placeholder="Type or select page number(s)...",
                    key=f"selected_pdf_pages_{st.session_state['of_reset_id']}"
                )
                selected_pdf_pages = sorted(int(page) for page in selected_pdf_pages)

            st.caption(
                "Page numbers are the actual PDF page numbers. They are never renumbered after selection."
            )

        # -----------------------------------------------------
        # PAGE → ORDER FORM ROW MAPPING
        # -----------------------------------------------------
        if selected_pdf_pages:
            st.markdown(
                '<div class="section-title">🔗 Page → Order Form Row Mapping</div>',
                unsafe_allow_html=True
            )

            if len(df) == 1:
                for page_number in selected_pdf_pages:
                    page_row_mapping[int(page_number)] = 0
                st.success(
                    "✅ Excel contains one data row. Every selected PDF page will use Data Row 1 (Excel Row 2)."
                )
                mapping_ready = True

            else:
                st.caption(
                    "Each selected PDF page is mapped to an Order Form data row. "
                    "The default follows Page N → Data Row N, but you can change it."
                )

                mapping_labels = [
                    f"Data Row {i + 1}  (Excel Row {i + 2})"
                    for i in range(len(df))
                ]

                all_mapped = True
                for page_number in selected_pdf_pages:
                    natural_index = int(page_number) - 1
                    default_index = natural_index if natural_index < len(df) else None

                    selected_row_label = st.selectbox(
                        f"PDF Page {page_number} → Order Form Data Row",
                        options=mapping_labels,
                        index=default_index,
                        placeholder="Select an Order Form data row...",
                        key=(
                            f"page_row_map_{page_number}_"
                            f"{st.session_state['of_reset_id']}"
                        )
                    )

                    if selected_row_label:
                        selected_row_index = mapping_labels.index(selected_row_label)
                        page_row_mapping[int(page_number)] = selected_row_index
                    else:
                        all_mapped = False

                mapping_ready = bool(selected_pdf_pages) and all_mapped

                if mapping_ready:
                    st.success("✅ Page-to-row mapping is ready for comparison.")
                else:
                    st.info(
                        "Please select an Order Form data row for every selected PDF page before comparison."
                    )

        # =========================================================
        # COMPARISON METHOD
        # =========================================================
        comparison_method = None
        selected_fields = []

        st.divider()
        st.markdown(
            '<div class="section-title">⚙️ Comparison Method</div>',
            unsafe_allow_html=True
        )
        st.caption(
            "Choose how the Order Form data should be matched to the Output."
        )

        comparison_method = st.radio(
            "Comparison Method",
            options=["Auto Detect", "Select Fields"],
            index=None,
            horizontal=True,
            key=f"comparison_method_{st.session_state['of_reset_id']}"
        )

        available_fields = get_available_fields(df)

        if comparison_method == "Auto Detect":
            auto_key = (
                AUTO_DETECT_ENGINE_VERSION,
                str(getattr(excel_file, "name", "")),
                int(getattr(excel_file, "size", 0)),
                str(getattr(output_file, "name", "")),
                int(getattr(output_file, "size", 0)),
                product_type,
                tuple(sorted(page_row_mapping.items()))
            )

            auto_widget_key = f"auto_selected_fields_{st.session_state['of_reset_id']}"

            if not mapping_ready:
                st.info(
                    "Complete the Page → Order Form Row Mapping first. Auto Detect will then read only the mapped artwork page(s)."
                )
                detected_fields = []
            else:
                if st.session_state.get("of_auto_detect_key") != auto_key:
                    try:
                        with st.spinner("Auto Detect is reading the mapped artwork page(s) with OCR..."):
                            auto_pages = extract_output_pages(output_file)
                            detected_fields = auto_detect_fields(
                                df,
                                auto_pages,
                                product_type,
                                page_row_mapping=page_row_mapping
                            )
                        st.session_state["of_auto_detected_fields"] = detected_fields
                        st.session_state["of_auto_output_pages"] = auto_pages
                        st.session_state["of_auto_detect_key"] = auto_key
                        # Reset the editable Auto Detect selection only when the
                        # files/mapping actually change. Manual edits then remain
                        # untouched on subsequent Streamlit reruns.
                        st.session_state[auto_widget_key] = list(detected_fields)
                    except Exception as error:
                        st.error(
                            f"Unable to run Auto Detect: {type(error).__name__}: {error}"
                        )
                        with st.expander("Technical error details", expanded=False):
                            import traceback
                            st.code(traceback.format_exc())

                detected_fields = st.session_state.get("of_auto_detected_fields", [])

            st.markdown(
                '<div class="section-title">🤖 Auto Detected Fields</div>',
                unsafe_allow_html=True
            )
            st.caption(
                "Only fields with strong evidence in the mapped artwork are pre-selected. "
                "You can remove a detected field or add another populated field before comparison."
            )

            default_auto = [field for field in detected_fields if field in available_fields]
            # Session state is populated above only when Auto Detect input changes.
            # This keeps the field list editable without Streamlit overwriting the
            # user's add/remove choices on every rerun.
            selected_fields = st.multiselect(
                "Review detected fields",
                options=available_fields,
                placeholder="Type to search or add a populated field...",
                label_visibility="collapsed",
                key=auto_widget_key
            )

            st.caption(
                f"Auto Detect engine: {AUTO_DETECT_ENGINE_VERSION} • "
                f"{len(selected_fields)} field(s) currently selected"
            )

            if selected_fields:
                st.caption("Selected fields: " + ", ".join(selected_fields))
            else:
                st.warning(
                    "Auto Detect did not find high-confidence populated fields in the mapped artwork. "
                    "You can add fields directly from the dropdown."
                )

        elif comparison_method == "Select Fields":
            st.markdown(
                '<div class="section-title">Select Variable Fields to Validate</div>',
                unsafe_allow_html=True
            )
            st.caption(
                "Only populated Order Form fields are shown. Search directly inside the dropdown by typing the field name."
            )

            previous = st.session_state.get(
                f"selected_fields_{st.session_state['of_reset_id']}",
                []
            )
            previous = [field for field in previous if field in available_fields]

            selected_fields = st.multiselect(
                "Select the fields from your Order Form",
                options=available_fields,
                default=previous,
                placeholder="Type to search fields...",
                label_visibility="collapsed",
                key=f"selected_fields_{st.session_state['of_reset_id']}"
            )

            if selected_fields:
                preview_rows = []
                for field in selected_fields:
                    values = []
                    for value in df[field].tolist():
                        if is_blank_value(value):
                            continue
                        values.append(str(value).strip())
                    preview_rows.append({
                        "Excel Field": field,
                        "Values": len(values),
                        "Preview": " | ".join(values[:3])
                    })

                with st.expander("🔎 Preview Selected Fields"):
                    st.dataframe(
                        pd.DataFrame(preview_rows),
                        width="stretch",
                        hide_index=True
                    )
            else:
                st.info(
                    "No fields are currently selected. Click the field dropdown and type to search for a populated field."
                )

        # =========================================================
        # COMPARE BUTTON
        # =========================================================
        st.markdown("<br>", unsafe_allow_html=True)

        compare_ready = (
            comparison_method is not None
            and bool(selected_fields)
            and product_type != "----- SELECT -----"
            and mapping_ready
        )

        if st.button(
            "🔍  COMPARE & PROOFREAD",
            width="stretch",
            key=f"of_compare_button_{st.session_state['of_reset_id']}",
            disabled=not compare_ready
        ):
            try:
                with st.spinner("Reading output and preparing comparison..."):
                    if (
                        comparison_method == "Auto Detect"
                        and st.session_state.get("of_auto_output_pages")
                    ):
                        output_pages = st.session_state["of_auto_output_pages"]
                    else:
                        output_pages = extract_output_pages(output_file)

                    if not output_pages:
                        raise ValueError("No readable output pages were detected.")

                    if not selected_fields:
                        raise ValueError("No fields are selected for comparison.")

                    st.session_state["of_report"] = build_report(
                        df,
                        output_pages,
                        selected_fields,
                        product_type,
                        page_row_mapping=page_row_mapping
                    )
                    st.session_state["of_report_selected_fields"] = selected_fields
                    st.session_state["of_report_product_type"] = product_type
                    st.session_state["of_report_comparison_method"] = comparison_method
                    st.session_state["of_report_build_version"] = AUTO_DETECT_ENGINE_VERSION
                    st.session_state["of_visual_pages"] = [
                        page for page in output_pages
                        if int(page.get("page", 0)) in selected_pdf_pages
                    ]

            except Exception as error:
                import traceback
                st.error(
                    f"Unable to process the Output Artwork: {type(error).__name__}: {error}"
                )
                with st.expander("Technical error details", expanded=True):
                    st.code(traceback.format_exc())
    # =========================================================
    # SAVED REPORT
    # =========================================================
    report = st.session_state.get("of_report")

    if report is not None:
        st.divider()

        st.markdown(
            '<div class="section-title">QC Report</div>',
            unsafe_allow_html=True
        )

        report_product_type = st.session_state.get(
            "of_report_product_type",
            product_type
        )

        report_method = st.session_state.get(
            "of_report_comparison_method",
            "Select Fields"
        )

        report_fields = st.session_state.get(
            "of_report_selected_fields",
            []
        )

        st.caption(
            f"Comparison method: {report_method} • "
            f"Product type: {report_product_type}"
        )

        if report_method == "Auto Detect" and report_fields:
            st.caption(
                "Auto-detected fields: " + ", ".join(report_fields)
            )

        pass_count = int(
            (report["STATUS"] == "PASS").sum()
        )

        fail_count = int(
            (report["STATUS"] == "FAIL").sum()
        )

        not_found_count = int(
            (report["STATUS"] == "NOT FOUND").sum()
        )

        skip_count = int(
            (report["STATUS"] == "SKIP").sum()
        )

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("PASS", pass_count)

        with col2:
            st.metric("FAIL", fail_count)

        with col3:
            st.metric("NOT FOUND", not_found_count)

        with col4:
            st.metric("IGNORED", skip_count)

        display_report = add_visual_column(report, report_fields)
        field_colors = get_field_visual_colors(report_fields)

        styled_report = (
            display_report
            .style
            .map(
                style_status,
                subset=["STATUS"]
            )
        )

        visual_styles = pd.DataFrame(
            "",
            index=display_report.index,
            columns=display_report.columns,
        )
        for row_index, field_name in display_report["FIELD"].items():
            if field_name in field_colors:
                visual_styles.loc[row_index, "VISUAL"] = (
                    f"background-color: {field_colors[field_name]};"
                    "color: white;"
                    "font-size: 17px;"
                    "font-weight: bold;"
                    "text-align: center;"
                )

        styled_report = styled_report.apply(
            lambda _df: visual_styles,
            axis=None,
        )

        st.dataframe(
            styled_report,
            width="stretch",
            hide_index=True
        )

        st.divider()

        # =========================================================
        # VISUAL ARTWORK COMPARISON
        # =========================================================
        visual_pages = st.session_state.get("of_visual_pages") or []
        if visual_pages:
            st.markdown(
                '<div class="section-title">🖼️ Visual Artwork Comparison</div>',
                unsafe_allow_html=True
            )
            st.caption(
                "Full artwork pages are shown below. OCR-detected comparison text is highlighted directly on the original artwork."
            )
            for page in visual_pages:
                page_num = page.get("page")
                page_report = report[report["PDF PAGE"] == page_num] if "PDF PAGE" in report.columns else report.iloc[0:0]
                highlighted = build_highlighted_page_image(
                    page,
                    page_report,
                    field_colors=field_colors,
                )
                if highlighted is not None:
                    st.markdown(f"**Artwork Page {page_num}**")
                    st.image(highlighted, width=620)

        with st.expander("ℹ️ How this validation works"):
            st.write(
                """
                **Variable-data validation**

                Only the fields selected from the Order Form are treated as variable artwork data.

                **OCR validation**

                Artwork pages are rendered as images and OCR is used as the primary extraction source for non-editable artwork.

                **Combined output lines**

                Multiple Order Form fields can be matched independently when the artwork prints them on one line.

                **Page mapping**

                Each selected PDF page is mapped to a specific Order Form data row.

                By default, Page N → Data Row N (Excel Row N+1), but the mapping can be changed manually.

                When the Excel contains only one data row, every selected PDF page uses Data Row 1 (Excel Row 2).

                **PFL mode**

                Panel-numbered artwork is treated as a continuous stream so selected variable data can continue from one panel into the next panel.

                **Mismatch detection**

                If the selected Order Form value is present in the PDF → PASS.

                If the expected value is absent but a relevant alternative value is detected → FAIL.

                If an Order Form field is blank, that field is not required and is ignored.
                """
            )

        excel_data = create_excel_report(
            report=report,
            product_type=report_product_type,
            comparison_method=report_method,
            selected_fields=report_fields,
            visual_pages=visual_pages
        )

        st.download_button(
            label="⬇️ Download Excel QC Report",
            data=excel_data,
            file_name="PDF_Proofreading_QC_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            key=f"of_download_excel_qc_report_{st.session_state['of_reset_id']}"
        )

    # =========================================================
    # INITIAL INSTRUCTIONS
    # =========================================================
    if not excel_file:
        st.caption("Upload an Order Form to begin.")
    elif not output_file:
        st.caption("Upload the Output Artwork to continue.")
    elif product_type == "----- SELECT -----":
        st.caption("Select the Product Type to continue.")
    elif comparison_method is None:
        st.caption("Select a Comparison Method to continue.")
    elif comparison_method == "Select Fields" and not selected_fields:
        st.caption("Select the variable fields you want to validate.")
